#!/usr/bin/env python3
"""
Script final optimisé pour importer DIATA ZONE B.xlsx
Gère spécifiquement chaque feuille avec sa structure connue
"""

import os
import sys
import re
import django
from datetime import datetime, date, timedelta
from dateutil.parser import parse as parse_date

# Setup Django
sys.path.append('/Users/omerlinks/Desktop/project/CLEAN')
os.environ.setdefault('DJANGO_SETTINGS_MODULE', 'core.settings')
django.setup()

import openpyxl
from business.models import Zone, Quartier, SubscriptionPlan, Subscription
from accounts.models import User
from django.db.models import Q

def parse_date(value):
    if value is None:
        return None
    if isinstance(value, (datetime, date)):
        return value.date() if isinstance(value, datetime) else value
    if isinstance(value, str):
        try:
            date_str = value.strip().lower()
            if date_str.startswith('le '):
                date_str = date_str[3:]
            return parse_date(date_str).date()
        except:
            pass
    return None

def clean_phone(phone):
    if phone is None:
        return None
    phone_str = str(phone).strip()
    if phone_str in ['////', '///', '//', '/', 'None', '']:
        return None
    cleaned = ''.join(c for c in phone_str if c.isdigit())
    return cleaned if cleaned and len(cleaned) >= 8 else None

def clean_name(name):
    if name is None:
        return None
    name_str = str(name).strip()
    if not name_str:
        return None
    return ' '.join(name_str.split())

def get_or_create_zone(name):
    try:
        return Zone.objects.get(name__iexact=name.strip())
    except Zone.DoesNotExist:
        return Zone.objects.create(name=name.strip())

def get_or_create_quartier(name, zone):
    name_clean = clean_name(name)
    if not name_clean:
        return None
    try:
        return Quartier.objects.get(name__iexact=name_clean, zone=zone)
    except Quartier.DoesNotExist:
        return Quartier.objects.create(name=name_clean, zone=zone)

def get_or_create_plan(name, price):
    name_clean = clean_name(name)
    if not name_clean:
        name_clean = "Bsic"
    
    name_lower = name_clean.lower()
    if 'basic' in name_lower or 'bsic' in name_lower:
        normalized = 'Bsic'
    elif 'standard' in name_lower:
        normalized = 'Standard'
    elif 'entreprise' in name_lower:
        normalized = 'Entreprise'
    elif '3 plus' in name_lower or 'plus' in name_lower:
        normalized = '3 Plus'
    else:
        normalized = name_clean
    
    try:
        return SubscriptionPlan.objects.get(name__iexact=normalized)
    except SubscriptionPlan.DoesNotExist:
        return SubscriptionPlan.objects.create(
            name=normalized,
            price=price,
            duration_days=30,
            description=f"Plan {normalized} - {price} FCFA"
        )

def get_or_create_client(name, phone, phone2, quartier, address):
    name_clean = clean_name(name)
    phone_clean = clean_phone(phone) if phone else None
    phone2_clean = clean_phone(phone2) if phone2 else None
    
    if not name_clean:
        if phone_clean:
            name_clean = f"Client_{phone_clean[-4:]}"
        else:
            name_clean = "Client_Inconnu"
    
    existing = User.objects.filter(
        Q(phone_number=phone_clean) | Q(phone_number_2=phone_clean),
        role=User.Role.CLIENT
    ).first()
    
    if existing:
        if phone2_clean and not existing.phone_number_2:
            existing.phone_number_2 = phone2_clean
        if quartier and not existing.quartier:
            existing.quartier = quartier
        if address and not existing.address:
            existing.address = address
        existing.save()
        return existing
    
    existing = User.objects.filter(
        username__iexact=name_clean,
        role=User.Role.CLIENT
    ).first()
    
    if existing:
        if phone_clean and not existing.phone_number:
            existing.phone_number = phone_clean
        if phone2_clean and not existing.phone_number_2:
            existing.phone_number_2 = phone2_clean
        if quartier and not existing.quartier:
            existing.quartier = quartier
        if address and not existing.address:
            existing.address = address
        existing.save()
        return existing
    
    return User.objects.create_user(
        username=name_clean,
        phone_number=phone_clean,
        phone_number_2=phone2_clean,
        quartier=quartier,
        address=address,
        role=User.Role.CLIENT
    )

def is_valid_name(name):
    """Vérifier si un nom est valide (pas une adresse, pas un nombre, etc.)"""
    if not name or len(name) < 2:
        return False
    
    name_lower = name.lower()
    
    # Noms invalides (adresses, etc.)
    invalid = ['diata', 'avenue', 'rue', 'av ', 'massangui', 'boupanda', 
              'mbila', 'mont fouary', 'makabana', 'mbongui', 'makola', 'yamba',
              'fevrier', 'chateau', 'libis', 'apoen', 'baconbgo',
              'ecole', 'eglise', 'residence', 'sum', '=sum']
    
    if any(inv in name_lower for inv in invalid):
        return False
    
    # Si le nom contient uniquement des chiffres ou des lettres sans espace
    if name.isdigit() or (len(name) < 3 and not any(c.isalpha() for c in name)):
        return False
    
    return True

def process_feuil1(sheet, zone):
    """Traiter Feuil1 avec sa structure spécifique"""
    created = 0
    errors = 0
    skipped = 0
    
    for row_idx, row in enumerate(sheet.iter_rows(), 1):
        try:
            values = [cell.value for cell in row]
            
            # Sauter les lignes vides
            if all(v is None or (isinstance(v, str) and not v.strip()) for v in values):
                continue
            
            # Sauter les en-têtes
            if any(isinstance(v, str) and v.strip().lower() in ['n', 'ruelle', 'quartier', 'nom', 'tel', 'forfait', 'pack', 'prix', 'date', 'dte', 'colonne'] for v in values):
                continue
            
            if any(isinstance(v, str) and 'diata zone b' in v.lower() for v in values):
                continue
            
            # Structure Feuil1: [N, rue_num, rue, quartier, nom, tel, tel2, forfait, prix, devise, date_debut, jour, date_recouv, ...]
            if len(values) < 13:
                skipped += 1
                continue
            
            nom = clean_name(values[4])  # Colonne 4 = Nom
            tel = clean_phone(values[5])  # Colonne 5 = Tel
            tel2 = clean_phone(values[6])  # Colonne 6 = Tel2
            forfait = clean_name(values[7])  # Colonne 7 = Forfait
            prix = values[8] if isinstance(values[8], (int, float)) and values[8] > 0 else None  # Colonne 8 = Prix
            date_debut_str = values[10]  # Colonne 10 = Date début
            date_recouv_str = values[12]  # Colonne 12 = Date recouvrement
            rue = clean_name(values[2])  # Colonne 2 = Rue
            
            # Si prix est None ou 0, essayer de trouver le prix dans le forfait d'abord
            if prix is None or prix == 0:
                # Vérifier si le forfait contient un nombre
                if forfait:
                    matches = re.findall(r'\d+', forfait)
                    if matches:
                        prix = float(matches[-1])  # Prendre le dernier nombre
                
                # Si toujours pas de prix, utiliser le prix par défaut selon le forfait
                if prix is None or prix == 0:
                    if forfait:
                        forfait_lower = forfait.lower()
                        if 'entreprise' in forfait_lower:
                            prix = 10000
                        elif 'standard' in forfait_lower:
                            prix = 5000
                        elif '3 plus' in forfait_lower or 'plus' in forfait_lower:
                            prix = 3500
                        elif 'bsic' in forfait_lower or 'basic' in forfait_lower:
                            prix = 2000
                        else:
                            prix = 2000  # Default
                
                # Sinon, essayer de trouver une valeur numérique dans les colonnes après 7
                if prix is None or prix == 0:
                    for i in range(8, min(len(values), 15)):
                        val = values[i]
                        if isinstance(val, (int, float)) and val > 0:
                            prix = float(val)
                            break
            
            if prix is None:
                prix = 2000
            
            if not forfait:
                if prix >= 10000:
                    forfait = "Entreprise"
                elif prix >= 5000:
                    forfait = "Standard"
                elif prix >= 2500:
                    forfait = "Standard"
                else:
                    forfait = "Bsic"
            
            # Vérifier si le nom est valide
            if not is_valid_name(nom):
                skipped += 1
                continue
            
            # Obtenir quartier
            quartier_name = clean_name(values[3]) if len(values) > 3 else "DIATA"
            if quartier_name and quartier_name.strip().lower() != 'diata':
                quartier = get_or_create_quartier(quartier_name, zone)
            else:
                quartier = get_or_create_quartier("DIATA", zone)
            
            # Obtenir plan
            plan = get_or_create_plan(forfait, prix)
            
            # Obtenir client
            client = get_or_create_client(nom, tel, tel2, quartier, rue)
            
            # Parser dates
            date_debut = parse_date(date_debut_str) if date_debut_str else date.today()
            if date_debut is None:
                date_debut = date.today()
            date_fin = parse_date(date_recouv_str) if date_recouv_str else None
            
            if date_fin is None:
                if date_debut is not None:
                    date_fin = date_debut + timedelta(days=30)
                else:
                    date_fin = date.today() + timedelta(days=30)
            
            # Créer abonnement
            Subscription.objects.create(
                client=client,
                plan=plan,
                start_date=date_debut,
                end_date=date_fin,
                is_active=True
            )
            
            created += 1
            # Debug: si c'est Bénie love, afficher plus d'info
            if nom and ('Bénie' in nom or 'love' in nom.lower()):
                print(f"  Created (DEBUG): {nom} | {plan.name} ({prix:.0f} FCFA) | {date_debut} to {date_fin} | Tel: {tel} | Tel2: {tel2} | Rue: {rue} | Quartier: {quartier_name}")
            else:
                print(f"  Created: {nom} | {plan.name} ({prix:.0f} FCFA) | {date_debut} to {date_fin}")
            
        except Exception as e:
            errors += 1
            print(f"  Error row {row_idx}: {str(e)}")
            continue
    
    return created, errors, skipped

def process_feuil2(sheet, zone):
    """Traiter Feuil2 avec sa structure spécifique"""
    created = 0
    errors = 0
    skipped = 0
    
    for row_idx, row in enumerate(sheet.iter_rows(), 1):
        try:
            values = [cell.value for cell in row]
            
            # Sauter les lignes vides
            if all(v is None or (isinstance(v, str) and not v.strip()) for v in values):
                continue
            
            # Sauter les en-têtes
            if any(isinstance(v, str) and v.strip().lower() in ['n', 'ruelle', 'quartier', 'nom', 'tel', 'forfait', 'pack', 'prix', 'date', 'dte', 'colonne'] for v in values):
                continue
            
            if any(isinstance(v, str) and 'diata zone b' in v.lower() for v in values):
                continue
            
            # Structure Feuil2: [None, None, N, rue, quartier, nom, tel, tel2, forfait, prix, devise, date_debut, ..., date_recouv]
            if len(values) < 14:
                skipped += 1
                continue
            
            # Les données commencent à la colonne 2
            nom = clean_name(values[5])  # Colonne 5 = Nom
            tel = clean_phone(values[6])  # Colonne 6 = Tel
            tel2 = clean_phone(values[7])  # Colonne 7 = Tel2
            forfait = clean_name(values[8])  # Colonne 8 = Forfait
            prix = values[9] if isinstance(values[9], (int, float)) and values[9] > 0 else None  # Colonne 9 = Prix
            date_debut_str = values[11]  # Colonne 11 = Date début
            date_recouv_str = values[13]  # Colonne 13 = Date recouvrement
            rue = clean_name(values[3])  # Colonne 3 = Rue
            
            # Si prix est None ou 0, essayer de trouver le prix dans le forfait d'abord
            if prix is None or prix == 0:
                # Vérifier si le forfait contient un nombre
                if forfait:
                    matches = re.findall(r'\d+', forfait)
                    if matches:
                        prix = float(matches[-1])  # Prendre le dernier nombre
                
                # Si toujours pas de prix, utiliser le prix par défaut selon le forfait
                if prix is None or prix == 0:
                    if forfait:
                        forfait_lower = forfait.lower()
                        if 'entreprise' in forfait_lower:
                            prix = 10000
                        elif 'standard' in forfait_lower:
                            prix = 5000
                        elif '3 plus' in forfait_lower or 'plus' in forfait_lower:
                            prix = 3500
                        elif 'bsic' in forfait_lower or 'basic' in forfait_lower:
                            prix = 2000
                        else:
                            prix = 2000  # Default
                
                # Sinon, essayer de trouver une valeur numérique dans les colonnes après 8
                if prix is None or prix == 0:
                    for i in range(9, min(len(values), 15)):
                        val = values[i]
                        if isinstance(val, (int, float)) and val > 0:
                            prix = float(val)
                            break
            
            if prix is None:
                prix = 2000
            
            if not forfait:
                if prix >= 10000:
                    forfait = "Entreprise"
                elif prix >= 5000:
                    forfait = "Standard"
                elif prix >= 2500:
                    forfait = "Standard"
                else:
                    forfait = "Bsic"
            
            # Vérifier si le nom est valide
            if not is_valid_name(nom):
                skipped += 1
                continue
            
            # Obtenir quartier
            quartier_name = clean_name(values[4]) if len(values) > 4 else "DIATA"
            if quartier_name and quartier_name.strip().lower() != 'diata':
                quartier = get_or_create_quartier(quartier_name, zone)
            else:
                quartier = get_or_create_quartier("DIATA", zone)
            
            # Obtenir plan
            plan = get_or_create_plan(forfait, prix)
            
            # Obtenir client
            client = get_or_create_client(nom, tel, tel2, quartier, rue)
            
            # Parser dates
            date_debut = parse_date(date_debut_str) if date_debut_str else date.today()
            if date_debut is None:
                date_debut = date.today()
            date_fin = parse_date(date_recouv_str) if date_recouv_str else None
            
            if date_fin is None:
                if date_debut is not None:
                    date_fin = date_debut + timedelta(days=30)
                else:
                    date_fin = date.today() + timedelta(days=30)
            
            # Créer abonnement
            Subscription.objects.create(
                client=client,
                plan=plan,
                start_date=date_debut,
                end_date=date_fin,
                is_active=True
            )
            
            created += 1
            # Debug: si c'est Bénie love, afficher plus d'info
            if nom and ('Bénie' in nom or 'love' in nom.lower()):
                print(f"  Created (DEBUG): {nom} | {plan.name} ({prix:.0f} FCFA) | {date_debut} to {date_fin} | Tel: {tel} | Tel2: {tel2} | Rue: {rue} | Quartier: {quartier_name}")
            else:
                print(f"  Created: {nom} | {plan.name} ({prix:.0f} FCFA) | {date_debut} to {date_fin}")
            
        except Exception as e:
            errors += 1
            print(f"  Error row {row_idx}: {str(e)}")
            continue
    
    return created, errors, skipped

def main():
    excel_path = '/Users/omerlinks/Desktop/project/CLEAN/scripts/DIATA ZONE B.xlsx'
    
    if not os.path.exists(excel_path):
        print(f"Error: File not found: {excel_path}")
        sys.exit(1)
    
    print(f"Loading Excel file: {excel_path}")
    
    try:
        wb = openpyxl.load_workbook(excel_path)
        print(f"Sheets: {wb.sheetnames}")
        
        zone = get_or_create_zone("DIATA")
        print(f"Using zone: {zone.name}")
        
        total_created = 0
        total_errors = 0
        total_skipped = 0
        
        for sheet_name in wb.sheetnames:
            sheet = wb[sheet_name]
            print(f"\nProcessing {sheet_name}...")
            if sheet_name == 'Feuil2':
                created, errors, skipped = process_feuil2(sheet, zone)
            else:
                created, errors, skipped = process_feuil1(sheet, zone)
            total_created += created
            total_errors += errors
            total_skipped += skipped
        
        print(f"\n{'='*70}")
        print(f"IMPORT SUMMARY")
        print(f"{'='*70}")
        print(f"✓ Subscriptions created: {total_created}")
        print(f"✗ Errors: {total_errors}")
        print(f"⊘ Skipped rows: {total_skipped}")
        
        print(f"\n📊 Database Status:")
        print(f"  🏙️  Zones: {Zone.objects.count()}")
        print(f"  🏘️  Quartiers: {Quartier.objects.count()}")
        print(f"  📋 Subscription Plans: {SubscriptionPlan.objects.count()}")
        print(f"  👥 Clients: {User.objects.filter(role=User.Role.CLIENT).count()}")
        print(f"  📄 Subscriptions: {Subscription.objects.count()}")
        
        print(f"\n💰 Plans d'abonnement:")
        for plan in SubscriptionPlan.objects.all().order_by('price'):
            subs_count = Subscription.objects.filter(plan=plan).count()
            total_price = subs_count * float(plan.price)
            print(f"  • {plan.name}: {plan.price:.0f} FCFA ({subs_count} abonnements) → {total_price:.0f} FCFA")
        
        print(f"\n📈 Répartition par forfait:")
        for plan in SubscriptionPlan.objects.all().order_by('name'):
            subs = Subscription.objects.filter(plan=plan)
            if subs.count() > 0:
                percentage = (subs.count() / total_created * 100) if total_created > 0 else 0
                print(f"  {plan.name}: {subs.count()} abonnements ({percentage:.1f}%)")
        
        print(f"\n✅ Import terminé avec succès pour DIATA ZONE B !")
        
    except Exception as e:
        print(f"Error: {str(e)}")
        import traceback
        traceback.print_exc()
        sys.exit(1)

if __name__ == '__main__':
    main()
