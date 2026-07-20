#!/usr/bin/env python3
"""
Import de TOUTES les zones depuis les fichiers Excel, puis envoi immédiat
vers la base de données en ligne (cloud).

RÈGLE MÉTIER IMPORTANTE :
  Le NOM DU FICHIER EXCEL définit la ZONE. Chaque fichier (ex: "DIATA ZONE A.xlsx",
  "Batignolle.xlsx", "Plateau des 15 ans TOUT 1.xlsx") crée/récupère une Zone
  portant le nom dérivé du fichier, et TOUS les clients importés depuis ce
  fichier sont liés à cette zone. Les zones sont poussées vers le cloud AVANT
  les clients (dépendance), donc aucun client n'est orphelin côté cloud.

Étapes :
  1. Pour chaque fichier Excel du dossier scripts/ :
       a. Dériver le nom de la zone du nom de fichier.
       b. Créer/récupérer la Zone (et la marquer synced=False pour le push).
       c. Importer les quartiers, plans, clients et abonnements liés à cette zone.
  2. Push automatique de toute la base locale vers le cloud via le moteur de
     synchronisation (push récursif des dépendances : Zone -> Quartier/User ->
     Subscription/Invoice).

Pré-requis :
  - Le cloud doit être joignable (paramètres CLOUD_DB_* dans .env).
  - Les dépendances openpyxl installées.

Usage :
  python scripts/import_to_cloud.py                 # import + push
  python scripts/import_to_cloud.py --no-push        # import local seulement
  python scripts/import_to_cloud.py --dry-run        # import local en simulation
"""

import io
import os
import re
import sys
from pathlib import Path

# Fix Windows console encoding (cp1252 can't handle emojis)
if sys.platform == "win32":
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")
    sys.stderr = io.TextIOWrapper(sys.stderr.buffer, encoding="utf-8", errors="replace")

BASE_DIR = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(BASE_DIR))
os.environ.setdefault("DJANGO_SETTINGS_MODULE", "core.settings")

import django

django.setup()

from django.db.models import Q

from accounts.models import User
from business.models import Quartier, Subscription, SubscriptionPlan, Zone
from sync_engine.cron import SyncDataCronJob
from sync_engine.router import SyncRouter

SCRIPTS_DIR = BASE_DIR / "scripts"

# Mots-clés en-tête pour détecter les lignes d'en-tête
HEADER_KEYWORDS = [
    "n",
    "ruelle",
    "quartier",
    "nom",
    "tel",
    "forfait",
    "pack",
    "prix",
    "date",
    "dte",
    "colonne",
    "numéro",
    "rue",
    "prenom",
    "prénom",
    "telephone",
    "téléphone",
    "devise",
    "recouvrement",
    "nom et prénom",
    "nom et prenom",
    "nom & prénom",
]

INVALID_NAME_PATTERNS = [
    "avenue",
    "rue",
    "av ",
    "résidence",
    "residence",
    "ecole",
    "école",
    "eglise",
    "église",
    "sum",
    "=sum",
    "=somme",
    "total",
    "sous-total",
    "sous total",
    "diata zone",
    "chateau d'eau",
    "batignolle",
    "la poudrielle",
    "plateau des",
    "mont fouary",
]


# ──────────────────────────────────────────────
# Helpers
# ──────────────────────────────────────────────
def derive_zone_name(filename):
    """Dériver le nom de la ZONE à partir du NOM DU FICHIER EXCEL.
    DIATA ZONE A/B/C -> 'DIATA' (une seule zone regroupée)."""
    name = Path(filename).stem
    name = re.sub(r"\s+TOUT\s*\d*$", "", name, flags=re.IGNORECASE)
    name = name.strip().upper()
    if re.match(r"^DIATA\s+ZONE\s+[A-Z]$", name):
        return "DIATA"
    return name


def clean_phone(phone):
    if phone is None:
        return None
    s = str(phone).strip()
    if s in ["////", "///", "//", "/", "None", ""]:
        return None
    cleaned = "".join(c for c in s if c.isdigit() or c.isspace())
    return cleaned.strip() or None


def clean_name(name):
    if name is None:
        return None
    s = str(name).strip()
    if not s:
        return None
    return " ".join(s.split())


def is_valid_client_name(nom):
    if not nom:
        return False
    low = nom.lower()
    if any(p in low for p in INVALID_NAME_PATTERNS):
        return False
    if any(kw in low for kw in HEADER_KEYWORDS):
        return False
    return True


def get_or_create_zone(name):
    """Créer ou récupérer une ZONE par nom (dérivé du fichier)."""
    name = name.strip()
    zone, created = Zone.objects.get_or_create(
        name__iexact=name, defaults={"name": name, "description": f"Zone {name}"}
    )
    return zone, created


def get_or_create_quartier(name, zone):
    name_clean = clean_name(name)
    if not name_clean:
        return None
    q, _ = Quartier.objects.get_or_create(
        name__iexact=name_clean, zone=zone, defaults={"name": name_clean, "zone": zone}
    )
    return q


def normalize_plan_name(name):
    if not name:
        return "Bsic"
    low = name.lower().strip()
    if "basic" in low or "bsic" in low or "basique" in low:
        return "Bsic"
    elif "standard" in low:
        return "Standard"
    elif "entreprise" in low or "enterprise" in low:
        return "Entreprise"
    elif "3 plus" in low or "3plus" in low or "plus" in low:
        return "3 Plus"
    return name.strip()


def get_default_price(plan_name):
    return {"Bsic": 2000, "Standard": 5000, "Entreprise": 10000, "3 Plus": 3500}.get(
        plan_name, 2000
    )


def infer_plan_from_price(price):
    if price is None or price <= 0:
        return "Bsic"
    if price >= 10000:
        return "Entreprise"
    elif price >= 5000:
        return "Standard"
    elif price >= 3000:
        return "3 Plus"
    return "Bsic"


def get_or_create_plan(name, price):
    normalized = normalize_plan_name(name)
    if price is None or price <= 0:
        price = get_default_price(normalized)
    plan, _ = SubscriptionPlan.objects.get_or_create(
        name__iexact=normalized,
        defaults={
            "name": normalized,
            "price": price,
            "duration_days": 30,
            "description": f"Plan {normalized} - {price:.0f} FCFA",
        },
    )
    return plan


def get_or_create_client(name, phone, phone2, quartier, zone, address=None):
    """Créer ou récupérer un CLIENT lié à sa ZONE.
    Deux clients peuvent avoir le meme nom : on ne deduit JAMAIS par nom seul.
    La distinction reelle se fait par uuid (ID de sync). On deduit par
    telephone si disponible ; sinon on cree toujours un nouveau client.
    Le vrai nom est stocké dans display_name, username est généré à partir de l'uuid."""
    phone_clean = clean_phone(phone)
    phone2_clean = clean_phone(phone2)

    if not name:
        if phone_clean:
            name = f"Client_{phone_clean[-4:]}"
        else:
            return None

    if phone_clean:
        existing = User.objects.filter(
            Q(phone_number=phone_clean) | Q(phone_number_2=phone_clean),
            role=User.Role.CLIENT,
        ).first()
        if existing:
            return existing
    # Pas de dedoublonnage par nom : on cree un client par ligne.
    return User.objects.create_user(
        # username laisse vide -> genere automatiquement depuis l'uuid
        first_name=name,
        display_name=name,
        phone_number=phone_clean,
        phone_number_2=phone2_clean,
        quartier=quartier,
        zone=zone,
        address=address,
        role=User.Role.CLIENT,
    )


def safe_parse_date(value):
    if value is None:
        return None
    if hasattr(value, "date"):
        return value.date()
    if isinstance(value, str):
        s = value.strip()
        for fmt in ("%d/%m/%Y", "%Y-%m-%d", "%d-%m-%Y", "%d.%m.%Y"):
            try:
                from datetime import datetime

                return datetime.strptime(s, fmt).date()
            except ValueError:
                continue
    return None


def detect_column_layout(sheet):
    header_row = None
    col_map = {}
    for row_idx, row in enumerate(sheet.iter_rows(max_row=15), 1):
        values_str = [str(v).strip().lower() if v else "" for v in row]
        score = sum(
            1 for v in values_str if v and any(kw in v for kw in HEADER_KEYWORDS)
        )
        if score >= 3:
            header_row = row_idx
            for i, v in enumerate(values_str):
                if not v:
                    continue
                if any(k in v for k in ["nom et pr", "nom & pr", "nom et pr"]):
                    col_map["nom"] = i
                elif v in ["n", "n°", "no", "num", "numéro"] and "nom" not in col_map:
                    col_map["n"] = i
                elif any(k in v for k in ["ruelle", "rue"]) and "rue" not in col_map:
                    col_map["rue"] = i
                elif "quartier" in v and "quartier" not in col_map:
                    col_map["quartier"] = i
                elif (
                    any(k in v for k in ["nom", "prenom", "prénom"])
                    and "nom" not in col_map
                ):
                    col_map["nom"] = i
                elif any(k in v for k in ["tel", "téléphone", "telephone", "phone"]):
                    if "tel" not in col_map:
                        col_map["tel"] = i
                    elif "tel2" not in col_map:
                        col_map["tel2"] = i
                elif any(k in v for k in ["forfait", "pack", "formule"]):
                    col_map["forfait"] = i
                elif v in ["prix", "montant", "tarif"]:
                    col_map["prix"] = i
                elif any(
                    k in v
                    for k in [
                        "date début",
                        "date debut",
                        "dte début",
                        "dte debut",
                        "date dbt",
                    ]
                ):
                    col_map["date_debut"] = i
                elif any(
                    k in v
                    for k in [
                        "recouvrement",
                        "date fin",
                        "date recouv",
                        "échéance",
                        "echeance",
                    ]
                ):
                    col_map["date_fin"] = i
                elif "date" in v and "date_debut" not in col_map:
                    col_map["date_debut"] = i
            break
    if not header_row:
        return {
            "layout": "default",
            "data_start_row": 5,
            "n": 0,
            "rue": 2,
            "quartier": 3,
            "nom": 4,
            "tel": 5,
            "tel2": 6,
            "forfait": 7,
            "prix": 8,
            "date_debut": 10,
            "date_fin": 12,
        }
    return {"layout": "detected", "data_start_row": header_row + 1, **col_map}


def _get_val(values, layout, key):
    idx = layout.get(key)
    if idx is not None and idx < len(values):
        return values[idx]
    return None


def process_sheet(sheet, zone, default_quartier_name):
    """Importer une feuille en liant chaque client à sa ZONE."""
    from datetime import date, timedelta

    stats = {"created": 0, "errors": 0, "skipped": 0, "updated": 0}
    layout = detect_column_layout(sheet)
    data_start = layout.get("data_start_row", 5)

    for row_idx, row in enumerate(sheet.iter_rows(min_row=data_start), data_start):
        try:
            values = [cell.value for cell in row]
            if all(v is None or (isinstance(v, str) and not v.strip()) for v in values):
                continue
            if any(
                isinstance(v, str) and v.strip().lower() in HEADER_KEYWORDS
                for v in values
            ):
                continue

            nom = clean_name(_get_val(values, layout, "nom"))
            tel = _get_val(values, layout, "tel")
            tel2 = _get_val(values, layout, "tel2")
            forfait = clean_name(_get_val(values, layout, "forfait"))
            prix_raw = _get_val(values, layout, "prix")
            date_debut_raw = _get_val(values, layout, "date_debut")
            date_fin_raw = _get_val(values, layout, "date_fin")
            rue = clean_name(_get_val(values, layout, "rue"))
            quartier_name = clean_name(_get_val(values, layout, "quartier"))

            if not is_valid_client_name(nom):
                stats["skipped"] += 1
                continue

            prix = None
            if isinstance(prix_raw, (int, float)):
                prix = float(prix_raw)
            elif isinstance(prix_raw, str):
                digits = "".join(c for c in prix_raw if c.isdigit())
                prix = float(digits) if digits else None

            if not forfait:
                forfait = infer_plan_from_price(prix) if prix and prix > 0 else "Bsic"
            if prix is None or prix <= 0:
                prix = get_default_price(normalize_plan_name(forfait))

            quartier = get_or_create_quartier(
                quartier_name or default_quartier_name, zone
            )
            plan = get_or_create_plan(forfait, prix)
            client = get_or_create_client(nom, tel, tel2, quartier, zone, address=rue)
            if client is None:
                stats["skipped"] += 1
                continue

            date_debut = safe_parse_date(date_debut_raw) or date.today()
            date_fin = safe_parse_date(date_fin_raw) or (
                date_debut + timedelta(days=30)
            )

            sub, sub_created = Subscription.objects.get_or_create(
                client=client,
                plan=plan,
                defaults={"end_date": date_fin, "is_active": True},
            )
            if not sub_created:
                if date_fin and sub.end_date != date_fin:
                    sub.end_date = date_fin
                    sub.save()
                stats["updated"] += 1
            else:
                stats["created"] += 1
        except Exception as e:
            stats["errors"] += 1
            print(f"    ❌ Erreur ligne {row_idx}: {e}")
            continue
    return stats


def import_all_zones_to_local(dry_run=False):
    print("=" * 70)
    print("📥 ÉTAPE 1/2 — Import des fichiers Excel (zone = nom de fichier)")
    print("=" * 70)

    excel_files = sorted(SCRIPTS_DIR.glob("*.xlsx"))
    if not excel_files:
        print("⚠️  Aucun fichier Excel trouvé dans scripts/")
        return

    print(f"📑 {len(excel_files)} fichier(s) trouvé(s)\n")
    grand_total = {"created": 0, "errors": 0, "skipped": 0, "updated": 0}

    for excel_file in excel_files:
        zone_name = derive_zone_name(excel_file.name)
        print(f"\n{'─' * 70}")
        print(f"📂 {excel_file.name}  →  ZONE: {zone_name}")
        print(f"{'─' * 70}")

        if dry_run:
            print("   ⚠️  DRY RUN : aucune écriture")
            continue

        zone, zone_created = get_or_create_zone(zone_name)
        if zone_created:
            print(f"   ✨ Zone '{zone_name}' CRÉÉE")
        else:
            print(f"   ♻️  Zone '{zone_name}' existe déjà")

        try:
            import openpyxl

            wb = openpyxl.load_workbook(str(excel_file), data_only=True)
        except Exception as e:
            print(f"   ❌ Impossible d'ouvrir {excel_file.name}: {e}")
            continue

        file_total = {"created": 0, "errors": 0, "skipped": 0, "updated": 0}
        for sheet_name in wb.sheetnames:
            sheet = wb[sheet_name]
            stats = process_sheet(sheet, zone, default_quartier_name=zone_name)
            for k in file_total:
                file_total[k] += stats.get(k, 0)
        wb.close()

        print(
            f"   📊 Créés: {file_total['created']} | Erreurs: {file_total['errors']} | Ignorés: {file_total['skipped']}"
        )
        for k in grand_total:
            grand_total[k] += file_total[k]

    print(f"\n{'=' * 70}")
    print("📊 RÉSUMÉ IMPORT LOCAL")
    print(f"{'=' * 70}")
    print(f"  ✅ Abonnements créés: {grand_total['created']}")
    print(f"  ❌ Erreurs: {grand_total['errors']}")
    print(f"  ⏭️  Ignorés: {grand_total['skipped']}")
    print(f"  🏙️  Zones: {Zone.objects.count()}")
    print(f"  👥 Clients: {User.objects.filter(role=User.Role.CLIENT).count()}")
    print(f"  📄 Abonnements: {Subscription.objects.count()}")


def main():
    import argparse

    parser = argparse.ArgumentParser(
        description="Importe les Excel (zone=nom de fichier) puis pousse vers le cloud."
    )
    parser.add_argument(
        "--no-push",
        action="store_true",
        help="Importe en base locale sans pousser vers le cloud.",
    )
    parser.add_argument(
        "--dry-run",
        action="store_true",
        help="Import local en simulation (aucune écriture).",
    )
    args = parser.parse_args()

    import_all_zones_to_local(dry_run=args.dry_run)

    if args.dry_run:
        print("\n⚠️  DRY RUN : aucun envoi vers le cloud (simulation).")
        return
    if args.no_push:
        print("\n⏭️  Option --no-push : données conservées en base locale uniquement.")
        return

    print("\n" + "=" * 70)
    print("☁️  ÉTAPE 2/2 — Envoi vers la base en ligne (cloud)")
    print("=" * 70)

    router = SyncRouter()
    if not router._check_cloud_connectivity():
        print(
            "❌ Cloud injoignable. Les données restent en base locale (synced=False)."
        )
        print("   Relance le script quand la connexion est disponible pour pousser.")
        sys.exit(1)

    try:
        job = SyncDataCronJob()
        job.do()  # push (local -> cloud) puis pull (cloud -> local)
        # Vérification : les zones sont-elles bien sur le cloud ?
        zones_local = Zone.objects.count()
        zones_cloud = Zone.objects.using("cloud").count()
        clients_local = User.objects.filter(role=User.Role.CLIENT).count()
        clients_cloud = (
            User.objects.using("cloud").filter(role=User.Role.CLIENT).count()
        )
        print("\n✅ Données envoyées vers le cloud avec succès.")
        print(f"   🏙️  Zones        : local={zones_local}  →  cloud={zones_cloud}")
        print(f"   👥 Clients      : local={clients_local}  →  cloud={clients_cloud}")
        print("   L'application en ligne affichera les nouvelles données après sync.")
    except Exception as e:
        print(f"\n❌ Erreur lors de l'envoi vers le cloud : {e}")
        import traceback

        traceback.print_exc()
        print(
            "\nLes données sont en base locale (synced=False). Relance le script pour retenter le push."
        )
        sys.exit(1)


if __name__ == "__main__":
    main()
