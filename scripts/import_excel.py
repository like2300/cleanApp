# scripts/import_excel.py

import os
import django
from datetime import datetime, timedelta
from openpyxl import load_workbook

os.environ.setdefault("DJANGO_SETTINGS_MODULE", "core.settings")
django.setup()

from core.models import Zone, Quartier, SubscriptionPlan, Subscription
from accounts.models import User


EXCEL_FILE = "DIATA ZONE B.xlsx"

wb = load_workbook(EXCEL_FILE)
ws = wb.active

# Création de la zone
zone, _ = Zone.objects.get_or_create(
    name="DIATA ZONE B"
)

for row in ws.iter_rows(min_row=6, values_only=True):

    if not row[4]:
        continue

    quartier_name = str(row[3]).strip()
    fullname = str(row[4]).strip()

    numero = row[5]
    numero2 = row[6]

    forfait = str(row[7]).strip()
    prix = row[8]

    date_debut = row[10]

    quartier, _ = Quartier.objects.get_or_create(
        name=quartier_name,
        zone=zone
    )

    plan, _ = SubscriptionPlan.objects.get_or_create(
        name=forfait,
        defaults={
            "price": prix,
            "duration_days": 30
        }
    )

    names = fullname.split()

    first_name = names[0]
    last_name = " ".join(names[1:]) if len(names) > 1 else ""

    username = str(numero).strip() if numero else fullname.replace(" ", "").lower()

    user, created = User.objects.get_or_create(
        username=username,
        defaults={
            "first_name": first_name,
            "last_name": last_name,
            "phone_number": str(numero) if numero else None
        }
    )

    if created:
        print(f"Utilisateur créé : {fullname}")

    # Calcul de la date de fin
    if date_debut:
        if isinstance(date_debut, datetime):
            end_date = date_debut + timedelta(days=plan.duration_days)
        else:
            end_date = date_debut
    else:
        end_date = None

    Subscription.objects.get_or_create(
        client=user,
        plan=plan,
        defaults={
            "start_date": date_debut,
            "end_date": end_date,
            "is_active": True
        }
    )

print("Import terminé.")
