#!/usr/bin/env python3
"""
Script pour importer les abonnements depuis le fichier Excel DIATA ZONE B.xlsx
"""

import os
import sys
import django
from datetime import datetime, date
from dateutil.parser import parse as parse_date

# Setup Django
sys.path.append('/Users/omerlinks/Desktop/project/CLEAN')
os.environ.setdefault('DJANGO_SETTINGS_MODULE', 'core.settings')
django.setup()

import openpyxl
from business.models import Zone, Quartier, SubscriptionPlan, Subscription
from accounts.models import User

def parse_excel_date(excel_date):
    """Convertir une date Excel en date Python"""
    if excel_date is None:
        return None
    if isinstance(excel_date, datetime):
        return excel_date.date()
    if isinstance(excel_date, date):
        return excel_date
    try:
        # Essayer de parser une chaîne de date
        return parse_date(str(excel_date)).date()
    except:
        return None

def clean_phone_number(phone):
    """Nettoyer un numéro de téléphone"""
    if phone is None:
        return None
    phone_str = str(phone).strip()
    if phone_str in ['////', '///', '//', '/', 'None', '']:
        return None
    # Garder uniquement les chiffres et espaces
    cleaned = ''.join(c for c in phone_str if c.isdigit() or c.isspace())
    return cleaned.strip() if cleaned.strip() else None

def clean_name(name):
    """Nettoyer un nom"""
    if name is None:
        return None
    name_str = str(name).strip()
    if not name_str:
        return None
    # Remplacer les espaces multiples par un seul
    return ' '.join(name_str.split())

def get_or_create_zone(name):
    """Obtenir ou créer une zone"""
    name_lower = name.strip().lower()
    try:
        return Zone.objects.get(name__iexact=name)
    except Zone.DoesNotExist:
        # Vérifier si une zone similaire existe
        existing = Zone.objects.filter(name__iexact=name_lower)
        if existing.exists():
            return existing.first()
        return Zone.objects.create(name=name.strip())

def get_or_create_quartier(name, zone):
    """Obtenir ou créer un quartier"""
    name_clean = clean_name(name)
    if not name_clean:
        return None
    try:
        return Quartier.objects.get(name__iexact=name_clean, zone=zone)
    except Quartier.DoesNotExist:
        return Quartier.objects.create(name=name_clean, zone=zone)

def get_or_create_subscription_plan(name, price):
    """Obtenir ou créer un plan d'abonnement"""
    name_clean = clean_name(name)
    if not name_clean:
        return None
    
    # Normaliser le nom
    name_normalized = name_clean.strip().lower()
    
    try:
        return SubscriptionPlan.objects.get(name__iexact=name_clean)
    except SubscriptionPlan.DoesNotExist:
        # Vérifier si un plan avec un nom similaire existe
        existing = SubscriptionPlan.objects.filter(name__iexact=name_normalized)
        if existing.exists():
            return existing.first()
        
        return SubscriptionPlan.objects.create(
            name=name_clean,
            price=price,
            duration_days=30,  # Par défaut 30 jours
            description=f"Plan {name_clean} - {price} FCFA"
        )

def get_or_create_client(username, phone, phone2, quartier, address):
    """Obtenir ou créer un client"""
    # Normaliser le nom d'utilisateur
    username_clean = clean_name(username)
    if not username_clean:
        username_clean = f"client_{phone}" if phone else "unknown"
    
    # Vérifier si l'utilisateur existe déjà par téléphone
    existing_by_phone = User.objects.filter(
        phone_number=phone,
        role=User.Role.CLIENT
    ).first()
    
    if existing_by_phone:
        # Mettre à jour les informations
        if phone2 and not existing_by_phone.phone_number_2:
            existing_by_phone.phone_number_2 = phone2
        if quartier and not existing_by_phone.quartier:
            existing_by_phone.quartier = quartier
        if address and not existing_by_phone.address:
            existing_by_phone.address = address
        existing_by_phone.save()
        return existing_by_phone
    
    # Vérifier si l'utilisateur existe par nom d'utilisateur
    existing_by_username = User.objects.filter(
        username__iexact=username_clean,
        role=User.Role.CLIENT
    ).first()
    
    if existing_by_username:
        if phone and not existing_by_username.phone_number:
            existing_by_username.phone_number = phone
        if phone2 and not existing_by_username.phone_number_2:
            existing_by_username.phone_number_2 = phone2
        if quartier and not existing_by_username.quartier:
            existing_by_username.quartier = quartier
        if address and not existing_by_username.address:
            existing_by_username.address = address
        existing_by_username.save()
        return existing_by_username
    
    # Créer un nouveau client
    return User.objects.create_user(
        username=username_clean,
        phone_number=phone,
        phone_number_2=phone2,
        quartier=quartier,
        address=address,
        role=User.Role.CLIENT
    )

def detect_column_indices(sheet):
    """Détecter automatiquement les indices des colonnes dans une feuille Excel"""
    # Lire les premières lignes pour détecter les en-têtes
    header_keywords = {
        'n': ['n', 'num', 'numero'],
        'ruelle': ['ruelle', 'rue', 'avenue', 'av '],
        'quartier': ['quartier', 'zone'],
        'nom': ['nom', 'name', 'prenom', 'nom et prenom'],
        'tel': ['tel', 'telephone', 'numero', 'phone'],
        'tel2': ['numero n2', 'tel2', 'phone2', 'second'],
        'forfait': ['forfait', 'pack', 'plan', 'abonn'],
        'prix': ['prix', 'price', 'montant', 'tarif'],
        'date_debut': ['date', 'debut', 'start', 'dte du debut'],
        'date_fin': ['fin', 'end', 'recouv', 'dte du recrov', 'échéance']
    }
    
    # Compter les occurrences de chaque mot-clé dans chaque colonne
    col_matches = {col: {keyword: 0 for keyword in header_keywords} for col in range(20)}
    
    # Lire les 10 premières lignes
    for row in sheet.iter_rows(min_row=1, max_row=min(sheet.max_row, 10)):
        for col_idx, cell in enumerate(row):
            if cell.value:
                cell_str = str(cell.value).strip().lower()
                for keyword, patterns in header_keywords.items():
                    for pattern in patterns:
                        if pattern in cell_str:
                            col_matches[col_idx][keyword] += 1
    
    # Déterminer la meilleure colonne pour chaque champ
    column_mapping = {}
    for field, patterns in header_keywords.items():
        max_count = 0
        best_col = None
        for col_idx, matches in col_matches.items():
            count = sum(matches[field] for field in header_keywords)
            if count > max_count:
                max_count = count
                best_col = col_idx
        column_mapping[field] = best_col
    
    return column_mapping

def find_numeric_value(row_values):
    """Trouver la première valeur numérique dans une liste"""
    for val in row_values:
        if isinstance(val, (int, float)) and val > 0:
            return float(val)
    return None

def find_string_value(row_values, keywords):
    """Trouver la première chaîne qui correspond aux mots-clés"""
    for val in row_values:
        if val and isinstance(val, str):
            val_lower = val.strip().lower()
            for keyword in keywords:
                if keyword in val_lower:
                    return clean_name(val)
    return None

def find_phone_values(row_values):
    """Trouver toutes les valeurs qui ressemblent à des numéros de téléphone"""
    phones = []
    for val in row_values:
        if val and isinstance(val, str):
            cleaned = clean_phone_number(val)
            if cleaned and len(cleaned.replace(' ', '')) >= 8:  # Un numéro de téléphone a au moins 8 chiffres
                phones.append(cleaned)
    return phones[:2]  # Retourner les 2 premiers numéros

def process_excel_sheet(sheet, zone_name="DIATA"):
    """Traiter une feuille Excel"""
    print(f"Processing sheet: {sheet.title}")
    
    # Obtenir ou créer la zone
    zone = get_or_create_zone(zone_name)
    print(f"Zone: {zone.name}")
    
    created_subscriptions = 0
    errors = []
    
    # Détecter automatiquement les colonnes
    column_mapping = detect_column_indices(sheet)
    print(f"Detected column mapping: {column_mapping}")
    
    # Lire toutes les lignes
    for row_idx, row in enumerate(sheet.iter_rows(), 1):
        try:
            row_values = [cell.value for cell in row]
            
            # Sauter les lignes vides
            if all(v is None or str(v).strip() == '' for v in row_values):
                continue
            
            # Sauter les lignes qui semblent être des en-têtes
            header_keywords = ['n', 'ruelle', 'quartier', 'nom', 'tel', 'forfait', 'pack', 'prix', 'date', 'dte']
            if any(str(v).strip().lower() in header_keywords for v in row_values if v):
                continue
            
            # Extraire les valeurs en utilisant le mapping détecté
            n = clean_name(row_values[column_mapping.get('n', 0)]) if column_mapping.get('n') is not None and len(row_values) > column_mapping.get('n', 0) else None
            quartier_name = clean_name(row_values[column_mapping.get('quartier', 3)]) if column_mapping.get('quartier') is not None and len(row_values) > column_mapping.get('quartier', 3) else None
            nom = clean_name(row_values[column_mapping.get('nom', 4)]) if column_mapping.get('nom') is not None and len(row_values) > column_mapping.get('nom', 4) else None
            
            # Extraire les numéros de téléphone de manière intelligente
            phones = find_phone_values(row_values)
            tel = phones[0] if len(phones) > 0 else None
            tel2 = phones[1] if len(phones) > 1 else None
            
            # Extraire le forfait (rechercher des mots-clés)
            forfait = find_string_value(row_values, ['bsic', 'basic', 'standard', 'entreprise', 'forfait', 'pack', '3 plus'])
            
            # Extraire le prix (première valeur numérique)
            prix = find_numeric_value(row_values)
            
            # Extraire les dates
            date_debut_str = None
            date_recouvrement_str = None
            
            for idx, val in enumerate(row_values):
                if val and isinstance(val, (datetime, date)):
                    if idx <= 15:  # Les dates sont généralement dans les premières colonnes
                        if date_debut_str is None:
                            date_debut_str = val
                        elif date_recouvrement_str is None:
                            date_recouvrement_str = val
                elif val and isinstance(val, str) and ('le ' in val.lower() or 'janv' in val.lower() or 'fev' in val.lower() or '202' in val):
                    if date_debut_str is None:
                        date_debut_str = val
                    elif date_recouvrement_str is None:
                        date_recouvrement_str = val
            
            # Si on n'a pas de nom ou téléphone, sauter
            if not nom and not tel:
                continue
            
            # Si on n'a pas de nom ou téléphone, sauter
            if not nom and not tel:
                continue
            
            # Si prix est None, utiliser une valeur par défaut
            if prix is None:
                prix = 2000  # Prix par défaut
            
            # Déduire le forfait du prix si non spécifié
            if not forfait and prix > 0:
                if prix >= 10000:
                    forfait = "Entreprise"
                elif prix >= 5000:
                    forfait = "Standard"
                elif prix >= 2500:
                    forfait = "Standard"
                else:
                    forfait = "Bsic"
            elif not forfait:
                forfait = "Bsic"  # Forfait par défaut
            
            # Extraire la rue (chercher une valeur qui ressemble à une adresse)
            rue = None
            for val in row_values:
                if val and isinstance(val, str):
                    val_lower = val.strip().lower()
                    if any(keyword in val_lower for keyword in ['av ', 'rue ', 'avenue ', 'massangui', 'boupanda', 'fevrier', 'mbila', 'mont fouary', 'makabana', 'mbongui']):
                        rue = clean_name(val)
                        break
            
            # Obtenir ou créer le quartier
            if not quartier_name or quartier_name.strip().lower() == 'diata':
                quartier = get_or_create_quartier("DIATA", zone)
            else:
                quartier = get_or_create_quartier(quartier_name, zone)
            
            # Obtenir ou créer le plan d'abonnement
            plan = get_or_create_subscription_plan(forfait, prix)
            
            # Créer ou obtenir le client
            client = get_or_create_client(nom, tel, tel2, quartier, rue)
            
            # Parser les dates
            date_debut = parse_excel_date(date_debut_str) if date_debut_str else date.today()
            date_recouvrement = parse_excel_date(date_recouvrement_str) if date_recouvrement_str else None
            
            # Si date_recouvrement n'existe pas, utiliser date_debut + 30 jours
            if date_recouvrement is None:
                from datetime import timedelta
                date_recouvrement = date_debut + timedelta(days=30)
            
            # Créer l'abonnement
            subscription = Subscription.objects.create(
                client=client,
                plan=plan,
                start_date=date_debut,
                end_date=date_recouvrement,
                is_active=True
            )
            
            created_subscriptions += 1
            print(f"Created subscription {created_subscriptions}: {client.username} - {plan.name} ({prix} FCFA) from {date_debut} to {date_recouvrement}")
            
        except Exception as e:
            errors.append(f"Row {row_idx}: {str(e)}")
            print(f"Error on row {row_idx}: {str(e)}")
            import traceback
            traceback.print_exc()
            continue
    
    return created_subscriptions, errors

def main():
    """Fonction principale"""
    excel_path = '/Users/omerlinks/Desktop/project/CLEAN/scripts/DIATA ZONE B.xlsx'
    
    if not os.path.exists(excel_path):
        print(f"Error: File not found: {excel_path}")
        sys.exit(1)
    
    print(f"Loading Excel file: {excel_path}")
    
    try:
        wb = openpyxl.load_workbook(excel_path)
        print(f"Loaded workbook with sheets: {wb.sheetnames}")
        
        total_created = 0
        all_errors = []
        
        # Traiter chaque feuille
        for sheet_name in wb.sheetnames:
            sheet = wb[sheet_name]
            created, errors = process_excel_sheet(sheet)
            total_created += created
            all_errors.extend(errors)
        
        print(f"\n=== Summary ===")
        print(f"Total subscriptions created: {total_created}")
        if all_errors:
            print(f"Total errors: {len(all_errors)}")
            for error in all_errors[:10]:  # Affiche les 10 premières erreurs
                print(f"  - {error}")
        else:
            print("No errors encountered")
        
        print(f"\n=== Database Status ===")
        print(f"Total zones: {Zone.objects.count()}")
        print(f"Total quartiers: {Quartier.objects.count()}")
        print(f"Total subscription plans: {SubscriptionPlan.objects.count()}")
        print(f"Total clients: {User.objects.filter(role=User.Role.CLIENT).count()}")
        print(f"Total subscriptions: {Subscription.objects.count()}")
        
    except Exception as e:
        print(f"Error: {str(e)}")
        import traceback
        traceback.print_exc()
        sys.exit(1)

if __name__ == '__main__':
    main()
