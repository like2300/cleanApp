#!/usr/bin/env python3
"""
Script amélioré pour importer les abonnements depuis DIATA ZONE B.xlsx
"""

import os
import sys
import django
from datetime import datetime, date, timedelta
from dateutil.parser import parse as parse_date

# Setup Django
sys.path.append('/Users/omerlinks/Desktop/project/CLEAN')
os.environ.setdefault('DJANGO_SETTINGS_MODULE', 'core.settings')
django.setup()

import openpyxl
from business.models import Zone, Quartier, SubscriptionPlan, Subscription
from accounts.models import User

def parse_excel_date(value):
    """Convertir une valeur Excel en date Python"""
    if value is None:
        return None
    if isinstance(value, (datetime, date)):
        return value.date() if isinstance(value, datetime) else value
    if isinstance(value, str):
        try:
            # Essayer de parser la chaîne
            date_str = value.strip().lower()
            if date_str.startswith('le '):
                date_str = date_str[3:]
            return parse_date(date_str).date()
        except:
            pass
    return None

def clean_phone(phone):
    """Nettoyer un numéro de téléphone"""
    if phone is None:
        return None
    phone_str = str(phone).strip()
    if phone_str in ['////', '///', '//', '/', 'None', '']:
        return None
    # Garder uniquement les chiffres
    cleaned = ''.join(c for c in phone_str if c.isdigit())
    return cleaned if cleaned and len(cleaned) >= 8 else None

def clean_name(name):
    """Nettoyer un nom"""
    if name is None:
        return None
    name_str = str(name).strip()
    if not name_str or name_str in ['None', '']:
        return None
    return ' '.join(name_str.split())

def find_phones_in_row(row_values):
    """Trouver tous les numéros de téléphone dans une ligne"""
    phones = []
    for val in row_values:
        if val and isinstance(val, str):
            cleaned = clean_phone(val)
            if cleaned:
                phones.append(cleaned)
    return phones[:2]  # Retourner max 2 numéros

def find_numeric_in_row(row_values):
    """Trouver la première valeur numérique > 0"""
    for val in row_values:
        if isinstance(val, (int, float)) and val > 0:
            return float(val)
    return None

def find_string_matching(row_values, keywords):
    """Trouver une chaîne qui match un des mots-clés"""
    for val in row_values:
        if val and isinstance(val, str):
            val_lower = val.strip().lower()
            for keyword in keywords:
                if keyword in val_lower:
                    return clean_name(val)
    return None

def find_date_in_row(row_values):
    """Trouver une date dans une ligne"""
    dates = []
    for val in row_values:
        parsed = parse_excel_date(val)
        if parsed:
            dates.append(parsed)
    return dates[:2]  # Retourner max 2 dates

def get_or_create_zone(name):
    """Obtenir ou créer une zone"""
    try:
        return Zone.objects.get(name__iexact=name.strip())
    except Zone.DoesNotExist:
        return Zone.objects.create(name=name.strip())

def get_or_create_quartier(name, zone):
    """Obtenir ou créer un quartier"""
    name_clean = clean_name(name)
    if not name_clean:
        return None
    try:
        return Quartier.objects.get(name__iexact=name_clean, zone=zone)
    except Quartier.DoesNotExist:
        return Quartier.objects.create(name=name_clean, zone=zone)

def get_or_create_plan(name, price):
    """Obtenir ou créer un plan d'abonnement"""
    name_clean = clean_name(name)
    if not name_clean:
        name_clean = "Standard"
    
    # Normaliser les noms courants
    name_normalized = name_clean.strip().lower()
    if 'basic' in name_normalized:
        name_normalized = 'Bsic'
    elif 'standard' in name_normalized:
        name_normalized = 'Standard'
    elif 'entreprise' in name_normalized:
        name_normalized = 'Entreprise'
    elif '3 plus' in name_normalized:
        name_normalized = '3 Plus'
    
    try:
        return SubscriptionPlan.objects.get(name__iexact=name_normalized)
    except SubscriptionPlan.DoesNotExist:
        return SubscriptionPlan.objects.create(
            name=name_normalized,
            price=price,
            duration_days=30,
            description=f"Plan {name_normalized} - {price} FCFA"
        )

def get_or_create_client(name, phone, phone2, quartier, address):
    """Obtenir ou créer un client"""
    name_clean = clean_name(name)
    phone_clean = clean_phone(phone) if phone else None
    phone2_clean = clean_phone(phone2) if phone2 else None
    
    # Si pas de nom, utiliser le téléphone comme nom
    if not name_clean and phone_clean:
        name_clean = f"Client_{phone_clean[-4:]}"
    elif not name_clean:
        name_clean = "Client_Inconnu"
    
    # Vérifier si le client existe déjà par téléphone
    existing = User.objects.filter(
        Q(phone_number=phone_clean) | Q(phone_number_2=phone_clean),
        role=User.Role.CLIENT
    ).first()
    
    if existing:
        if phone2_clean and not existing.phone_number_2:
            existing.phone_number_2 = phone2_clean
        if quartier and not existing.quartier:
            existing.quartier = quartier
        if address and not existing.address:
            existing.address = address
        existing.save()
        return existing
    
    # Vérifier par nom d'utilisateur
    existing = User.objects.filter(
        username__iexact=name_clean,
        role=User.Role.CLIENT
    ).first()
    
    if existing:
        if phone_clean and not existing.phone_number:
            existing.phone_number = phone_clean
        if phone2_clean and not existing.phone_number_2:
            existing.phone_number_2 = phone2_clean
        if quartier and not existing.quartier:
            existing.quartier = quartier
        if address and not existing.address:
            existing.address = address
        existing.save()
        return existing
    
    # Créer un nouveau client
    return User.objects.create_user(
        username=name_clean,
        phone_number=phone_clean,
        phone_number_2=phone2_clean,
        quartier=quartier,
        address=address,
        role=User.Role.CLIENT
    )

def process_sheet(sheet, zone_name="DIATA"):
    """Traiter une feuille Excel"""
    print(f"\nProcessing sheet: {sheet.title}")
    
    zone = get_or_create_zone(zone_name)
    print(f"Zone: {zone.name}")
    
    created = 0
    errors = 0
    
    for row_idx, row in enumerate(sheet.iter_rows(), 1):
        try:
            values = [cell.value for cell in row]
            
            # Sauter les lignes vides
            if all(v is None or (isinstance(v, str) and not v.strip()) for v in values):
                continue
            
            # Sauter les lignes d'en-tête
            header_indicators = ['n', 'ruelle', 'quartier', 'nom', 'tel', 'forfait', 'pack', 'prix', 'date', 'dte', 'num']
            if any(isinstance(v, str) and v.strip().lower() in header_indicators for v in values):
                continue
            
            # Extraire les informations
            phones = find_phones_in_row(values)
            tel = phones[0] if phones else None
            tel2 = phones[1] if len(phones) > 1 else None
            
            # Trouver le nom
            name = None
            for val in values:
                if val and isinstance(val, str):
                    # Ne pas prendre les valeurs qui ressemblent à des numéros ou des adresses
                    val_lower = val.strip().lower()
                    if not any(kw in val_lower for kw in ['av ', 'rue ', 'avenue', 'massangui', 'boupanda', 'mbila', 'mont fouary', 'makabana', 'mbongui', 'fevrier', 'diata', 'zone']):
                        if val.strip() and not any(c.isdigit() for c in val.strip()):
                            name = clean_name(val)
                            break
            
            # Si pas de nom trouvé, prendre la première chaîne non vide
            if not name:
                for val in values:
                    if val and isinstance(val, str) and val.strip():
                        name = clean_name(val)
                        break
            
            # Trouver le prix
            prix = find_numeric_in_row(values)
            if prix is None:
                prix = 2000  # Prix par défaut
            
            # Trouver le forfait
            forfait = find_string_matching(values, ['bsic', 'basic', 'standard', 'entreprise', '3 plus', 'forfait', 'pack'])
            if not forfait:
                # Déduire du prix
                if prix >= 10000:
                    forfait = "Entreprise"
                elif prix >= 5000:
                    forfait = "Standard"
                elif prix >= 2500:
                    forfait = "Standard"
                else:
                    forfait = "Bsic"
            
            # Trouver les dates
            dates = find_date_in_row(values)
            date_debut = dates[0] if dates else date.today()
            date_fin = dates[1] if len(dates) > 1 else None
            
            # Trouver la rue/quartier
            quartier_name = "DIATA"
            rue = None
            for val in values:
                if val and isinstance(val, str):
                    val_lower = val.strip().lower()
                    if any(kw in val_lower for kw in ['av ', 'rue ', 'avenue', 'massangui', 'boupanda', 'mbila', 'mont fouary', 'makabana', 'mbongui']):
                        if 'diata' not in val_lower:
                            rue = clean_name(val)
            
            # Obtenir ou créer le quartier
            quartier = get_or_create_quartier(quartier_name, zone)
            
            # Obtenir ou créer le plan
            plan = get_or_create_plan(forfait, prix)
            
            # Obtenir ou créer le client
            client = get_or_create_client(name, tel, tel2, quartier, rue)
            
            # Si pas de date de fin, utiliser date_debut + 30 jours
            if date_fin is None:
                date_fin = date_debut + timedelta(days=30)
            
            # Créer l'abonnement
            subscription = Subscription.objects.create(
                client=client,
                plan=plan,
                start_date=date_debut,
                end_date=date_fin,
                is_active=True
            )
            
            created += 1
            print(f"  Created: {client.username} | {plan.name} ({prix} FCFA) | {date_debut} to {date_fin}")
            
        except Exception as e:
            errors += 1
            print(f"  Error row {row_idx}: {str(e)}")
            import traceback
            traceback.print_exc()
            continue
    
    return created, errors

def main():
    """Fonction principale"""
    excel_path = '/Users/omerlinks/Desktop/project/CLEAN/scripts/DIATA ZONE B.xlsx'
    
    if not os.path.exists(excel_path):
        print(f"Error: File not found: {excel_path}")
        sys.exit(1)
    
    print(f"Loading Excel file: {excel_path}")
    
    try:
        wb = openpyxl.load_workbook(excel_path)
        print(f"Sheets: {wb.sheetnames}")
        
        total_created = 0
        total_errors = 0
        
        for sheet_name in wb.sheetnames:
            sheet = wb[sheet_name]
            created, errors = process_sheet(sheet, zone_name="DIATA")
            total_created += created
            total_errors += errors
        
        print(f"\n=== Summary ===")
        print(f"Subscriptions created: {total_created}")
        print(f"Errors: {total_errors}")
        
        print(f"\n=== Database Status ===")
        print(f"Zones: {Zone.objects.count()}")
        print(f"Quartiers: {Quartier.objects.count()}")
        print(f"Subscription Plans: {SubscriptionPlan.objects.count()}")
        print(f"Clients: {User.objects.filter(role=User.Role.CLIENT).count()}")
        print(f"Subscriptions: {Subscription.objects.count()}")
        
        # Afficher les plans créés
        print(f"\nPlans d'abonnement:")
        for plan in SubscriptionPlan.objects.all():
            print(f"  - {plan.name}: {plan.price} FCFA ({plan.duration_days} jours)")
        
    except Exception as e:
        print(f"Error: {str(e)}")
        import traceback
        traceback.print_exc()
        sys.exit(1)

if __name__ == '__main__':
    from django.db.models import Q
    main()
