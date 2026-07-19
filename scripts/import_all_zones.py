#!/usr/bin/env python3
"""
Script universel d'import de TOUTES les zones depuis les fichiers Excel.
Détecte automatiquement les colonnes, crée les zones à partir des noms de fichiers,
et importe clients, quartiers, plans et abonnements.

Usage:
    python manage.py shell -c "exec(open('scripts/import_all_zones.py').read())"
    ou:
    python scripts/import_all_zones.py
"""

import os
import sys
import re
import glob
import io
from pathlib import Path
from datetime import datetime, date, timedelta

# Fix Windows console encoding (cp1252 can't handle emojis)
if sys.platform == 'win32':
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')
    sys.stderr = io.TextIOWrapper(sys.stderr.buffer, encoding='utf-8', errors='replace')

# ──────────────────────────────────────────────
# Django Setup
# ──────────────────────────────────────────────
BASE_DIR = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(BASE_DIR))
os.environ.setdefault('DJANGO_SETTINGS_MODULE', 'core.settings')

import django
django.setup()

import openpyxl
from django.db.models import Q
from business.models import Zone, Quartier, SubscriptionPlan, Subscription
from accounts.models import User

# ──────────────────────────────────────────────
# Configuration
# ──────────────────────────────────────────────
SCRIPTS_DIR = BASE_DIR / 'scripts'

# Mapping spécial nom de fichier → nom de zone (quand le nom de fichier
# ne correspond pas exactement au nom de zone souhaité)
ZONE_NAME_OVERRIDES = {
    # "Plateau des 15 ans TOUT 1.xlsx": "PLATEAU DES 15 ANS",
    # Ajouter d'autres overrides si nécessaire
}

# Mots-clés en-tête pour détecter les lignes d'en-tête
HEADER_KEYWORDS = [
    'n', 'ruelle', 'quartier', 'nom', 'tel', 'forfait', 'pack',
    'prix', 'date', 'dte', 'colonne', 'numéro', 'rue', 'prenom',
    'prénom', 'telephone', 'téléphone', 'devise', 'recouvrement',
    'nom et prénom', 'nom et prenom', 'nom & prénom',
]

# Noms invalides (adresses, lieux, formules, etc.)
INVALID_NAME_PATTERNS = [
    'avenue', 'rue', 'av ', 'résidence', 'residence', 'ecole', 'école',
    'eglise', 'église', 'sum', '=sum', '=somme', 'total', 'sous-total',
    'sous total', 'diata zone', 'chateau d\'eau', 'batignolle',
    'la poudrielle', 'plateau des', 'mont fouary',
]


# ──────────────────────────────────────────────
# Utility Functions
# ──────────────────────────────────────────────
def clean_phone(phone):
    """Nettoyer et valider un numéro de téléphone."""
    if phone is None:
        return None
    phone_str = str(phone).strip()
    if phone_str in ['////', '///', '//', '/', 'None', '', '-', '--', 'N/A', 'n/a']:
        return None
    # Enlever les espaces, tirets, points, etc.
    cleaned = ''.join(c for c in phone_str if c.isdigit())
    # Numéro valide: au moins 6 chiffres
    return cleaned if cleaned and len(cleaned) >= 6 else None


def clean_name(name):
    """Nettoyer un nom (supprimer espaces multiples, etc.)."""
    if name is None:
        return None
    name_str = str(name).strip()
    if not name_str or name_str.lower() in ['none', 'n/a', '-', '--', '/', '//']:
        return None
    return ' '.join(name_str.split())


def is_valid_client_name(name):
    """Vérifier si un nom est valide pour un client."""
    if not name or len(name) < 2:
        return False

    name_lower = name.lower().strip()

    # Rejeter les formules, totaux, etc.
    if name_lower.startswith('=') or name_lower.startswith('+'):
        return False

    # Rejeter les noms qui sont juste des chiffres
    if name.replace(' ', '').replace('-', '').isdigit():
        return False

    # Rejeter les noms trop courts sans lettres
    if len(name) < 3 and not any(c.isalpha() for c in name):
        return False

    # Rejeter les noms qui ressemblent à des adresses ou lieux
    for pattern in INVALID_NAME_PATTERNS:
        if pattern in name_lower:
            return False

    return True


def safe_parse_date(value):
    """Parser une date de manière sécurisée."""
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if isinstance(value, (int, float)):
        # Pourrait être un serial number Excel
        try:
            from datetime import timedelta
            # Excel date serial: nombre de jours depuis 1899-12-30
            if 40000 < value < 50000:  # Plage raisonnable pour dates récentes
                return date(1899, 12, 30) + timedelta(days=int(value))
        except:
            pass
        return None
    if isinstance(value, str):
        value = value.strip()
        if not value:
            return None
        # Nettoyer les préfixes courants
        for prefix in ['le ', 'Le ', 'LE ']:
            if value.startswith(prefix):
                value = value[len(prefix):]

        # Essayer plusieurs formats de date
        formats = [
            '%d/%m/%Y', '%d-%m-%Y', '%d.%m.%Y',
            '%Y-%m-%d', '%Y/%m/%d',
            '%d/%m/%y', '%d-%m-%y',
            '%d %B %Y', '%d %b %Y',
        ]
        for fmt in formats:
            try:
                return datetime.strptime(value, fmt).date()
            except ValueError:
                continue
        # Dernier recours: dateutil
        try:
            from dateutil.parser import parse as dateutil_parse
            return dateutil_parse(value, dayfirst=True).date()
        except:
            pass
    return None


def derive_zone_name(filename):
    """Dériver le nom de la zone à partir du nom du fichier Excel.
    Les fichiers DIATA ZONE A/B/C sont regroupés dans une seule zone DIATA."""
    if filename in ZONE_NAME_OVERRIDES:
        return ZONE_NAME_OVERRIDES[filename]

    # Enlever l'extension
    name = Path(filename).stem

    # Enlever les suffixes de type "TOUT 1", "TOUT 2", etc.
    name = re.sub(r'\s+TOUT\s*\d*$', '', name, flags=re.IGNORECASE)

    # Normaliser en majuscules
    name = name.strip().upper()

    # DIATA ZONE A, DIATA ZONE B, DIATA ZONE C → une seule zone "DIATA"
    if re.match(r'^DIATA\s+ZONE\s+[A-Z]$', name):
        return 'DIATA'

    return name


def derive_default_quartier(filename):
    """Dériver le nom du quartier par défaut à partir du nom du fichier.
    Pour les fichiers DIATA, le quartier est DIATA ZONE A/B/C.
    Pour les autres, le quartier est le nom de la zone."""
    # Enlever l'extension
    name = Path(filename).stem
    name = re.sub(r'\s+TOUT\s*\d*$', '', name, flags=re.IGNORECASE)
    name = name.strip().upper()
    return name


# ──────────────────────────────────────────────
# Database Helpers
# ──────────────────────────────────────────────
def get_or_create_zone(name):
    """Créer ou récupérer une zone par nom."""
    name = name.strip()
    try:
        zone = Zone.objects.get(name__iexact=name)
        return zone, False
    except Zone.DoesNotExist:
        zone = Zone.objects.create(name=name, description=f"Zone {name}")
        return zone, True


def get_or_create_quartier(name, zone):
    """Créer ou récupérer un quartier dans une zone."""
    name_clean = clean_name(name)
    if not name_clean:
        return None
    try:
        return Quartier.objects.get(name__iexact=name_clean, zone=zone)
    except Quartier.DoesNotExist:
        return Quartier.objects.create(name=name_clean, zone=zone)


def normalize_plan_name(name):
    """Normaliser le nom du forfait."""
    if not name:
        return "Bsic"
    name_lower = name.lower().strip()
    if 'basic' in name_lower or 'bsic' in name_lower or 'basique' in name_lower:
        return 'Bsic'
    elif 'standard' in name_lower:
        return 'Standard'
    elif 'entreprise' in name_lower or 'enterprise' in name_lower:
        return 'Entreprise'
    elif '3 plus' in name_lower or '3plus' in name_lower or 'plus' in name_lower:
        return '3 Plus'
    return name.strip()


def get_default_price(plan_name):
    """Obtenir le prix par défaut pour un plan."""
    prices = {
        'Bsic': 2000,
        'Standard': 5000,
        'Entreprise': 10000,
        '3 Plus': 3500,
    }
    return prices.get(plan_name, 2000)


def infer_plan_from_price(price):
    """Inférer le nom du plan à partir du prix."""
    if price is None or price <= 0:
        return 'Bsic'
    if price >= 10000:
        return 'Entreprise'
    elif price >= 5000:
        return 'Standard'
    elif price >= 3000:
        return '3 Plus'
    else:
        return 'Bsic'


def get_or_create_plan(name, price):
    """Créer ou récupérer un plan d'abonnement."""
    normalized = normalize_plan_name(name)

    if price is None or price <= 0:
        price = get_default_price(normalized)

    try:
        return SubscriptionPlan.objects.get(name__iexact=normalized)
    except SubscriptionPlan.DoesNotExist:
        return SubscriptionPlan.objects.create(
            name=normalized,
            price=price,
            duration_days=30,
            description=f"Plan {normalized} - {price:.0f} FCFA"
        )


def get_or_create_client(name, phone, phone2, quartier, zone, address=None):
    """
    Créer ou récupérer un client avec déduplication.
    Recherche d'abord par téléphone, puis par nom.
    """
    phone_clean = clean_phone(phone)
    phone2_clean = clean_phone(phone2)

    if not name:
        if phone_clean:
            name = f"Client_{phone_clean[-4:]}"
        else:
            return None  # On ne peut pas créer un client sans nom ni téléphone

    # 1) Chercher par numéro de téléphone
    if phone_clean:
        existing = User.objects.filter(
            Q(phone_number=phone_clean) | Q(phone_number_2=phone_clean),
            role=User.Role.CLIENT
        ).first()
        if existing:
            _update_client_fields(existing, phone_clean, phone2_clean, quartier, zone, address)
            return existing

    # 2) Chercher par nom (username)
    existing = User.objects.filter(
        username__iexact=name,
        role=User.Role.CLIENT
    ).first()
    if existing:
        _update_client_fields(existing, phone_clean, phone2_clean, quartier, zone, address)
        return existing

    # 3) Créer un nouveau client
    # Générer un username unique
    username = name
    counter = 1
    while User.objects.filter(username__iexact=username).exists():
        username = f"{name}_{counter}"
        counter += 1

    user = User.objects.create_user(
        username=username,
        phone_number=phone_clean,
        phone_number_2=phone2_clean,
        quartier=quartier,
        zone=zone,
        address=address,
        role=User.Role.CLIENT
    )
    return user


def _update_client_fields(user, phone, phone2, quartier, zone, address):
    """Mettre à jour les champs manquants d'un client existant."""
    changed = False
    if phone and not user.phone_number:
        user.phone_number = phone
        changed = True
    if phone2 and not user.phone_number_2:
        user.phone_number_2 = phone2
        changed = True
    if quartier and not user.quartier:
        user.quartier = quartier
        changed = True
    if zone and not user.zone:
        user.zone = zone
        changed = True
    if address and not user.address:
        user.address = address
        changed = True
    if changed:
        user.save()


# ──────────────────────────────────────────────
# Column Detection
# ──────────────────────────────────────────────
def detect_column_layout(sheet):
    """
    Détecte automatiquement la disposition des colonnes dans une feuille.
    Retourne un dict avec les indices de colonnes et la première ligne de données.
    """
    # Scanner les 15 premières lignes pour trouver les en-têtes
    header_row = None
    col_map = {}

    for row_idx, row in enumerate(sheet.iter_rows(max_row=15), 1):
        values = [cell.value for cell in row]
        values_str = [str(v).strip().lower() if v else '' for v in values]

        # Vérifier si c'est une ligne d'en-tête
        header_score = sum(1 for v in values_str if v and any(kw in v for kw in HEADER_KEYWORDS))
        if header_score >= 3:  # Au moins 3 mots-clés d'en-tête trouvés
            header_row = row_idx
            # Mapper les colonnes
            for i, v in enumerate(values_str):
                if not v:
                    continue
                if any(k in v for k in ['nom et pr', 'nom & pr', 'nom et pr']):
                    col_map['nom'] = i
                elif v in ['n', 'n°', 'no', 'num', 'numéro'] and 'nom' not in col_map:
                    col_map['n'] = i
                elif any(k in v for k in ['ruelle', 'rue']) and 'rue' not in col_map:
                    col_map['rue'] = i
                elif 'quartier' in v and 'quartier' not in col_map:
                    col_map['quartier'] = i
                elif any(k in v for k in ['nom', 'prenom', 'prénom']) and 'nom' not in col_map:
                    col_map['nom'] = i
                elif any(k in v for k in ['tel', 'téléphone', 'telephone', 'phone']):
                    if 'tel' not in col_map:
                        col_map['tel'] = i
                    elif 'tel2' not in col_map:
                        col_map['tel2'] = i
                elif any(k in v for k in ['forfait', 'pack', 'formule']):
                    col_map['forfait'] = i
                elif v in ['prix', 'montant', 'tarif']:
                    col_map['prix'] = i
                elif any(k in v for k in ['devise', 'fcfa', 'xaf']):
                    col_map['devise'] = i
                elif any(k in v for k in ['date début', 'date debut', 'dte début', 'dte debut', 'date dbt']):
                    col_map['date_debut'] = i
                elif any(k in v for k in ['recouvrement', 'date fin', 'date recouv', 'échéance', 'echeance']):
                    col_map['date_fin'] = i
                elif 'date' in v and 'date_debut' not in col_map:
                    col_map['date_debut'] = i
                elif 'jour' in v:
                    col_map['jour'] = i
            break

    if not header_row:
        # Pas d'en-tête trouvé, utiliser la disposition par défaut (Feuil1 style)
        return {
            'layout': 'default',
            'data_start_row': 5,
            'n': 0, 'rue': 2, 'quartier': 3, 'nom': 4,
            'tel': 5, 'tel2': 6, 'forfait': 7, 'prix': 8,
            'date_debut': 10, 'date_fin': 12,
        }

    return {
        'layout': 'detected',
        'data_start_row': header_row + 1,
        **col_map,
    }


# ──────────────────────────────────────────────
# Sheet Processing
# ──────────────────────────────────────────────
def process_sheet(sheet, zone, zone_name, sheet_name, default_quartier_name=None):
    """Traiter une feuille Excel et importer les données."""
    stats = {'created': 0, 'errors': 0, 'skipped': 0, 'updated': 0}
    # Le quartier par défaut est le nom dérivé du fichier (ex: DIATA ZONE A)
    fallback_quartier = default_quartier_name or zone_name

    # Détecter la disposition des colonnes
    layout = detect_column_layout(sheet)
    data_start = layout.get('data_start_row', 5)

    print(f"    📐 Layout détecté: {layout.get('layout', 'unknown')}")
    print(f"    📍 Données à partir de la ligne {data_start}")
    print(f"    📊 Colonnes: { {k: v for k, v in layout.items() if k not in ('layout', 'data_start_row')} }")

    for row_idx, row in enumerate(sheet.iter_rows(min_row=data_start), data_start):
        try:
            values = [cell.value for cell in row]

            # Sauter les lignes vides
            if all(v is None or (isinstance(v, str) and not v.strip()) for v in values):
                continue

            # Sauter les lignes qui ressemblent à des en-têtes
            if any(isinstance(v, str) and v.strip().lower() in HEADER_KEYWORDS for v in values):
                continue

            # Sauter les lignes qui contiennent le nom de la zone comme titre
            if any(isinstance(v, str) and zone_name.lower() in v.lower() for v in values):
                continue

            # Extraire les données selon le layout détecté
            nom = _get_val(values, layout, 'nom')
            tel = _get_val(values, layout, 'tel')
            tel2 = _get_val(values, layout, 'tel2')
            forfait = _get_val(values, layout, 'forfait')
            prix_raw = _get_val(values, layout, 'prix')
            date_debut_raw = _get_val(values, layout, 'date_debut')
            date_fin_raw = _get_val(values, layout, 'date_fin')
            rue = _get_val(values, layout, 'rue')
            quartier_name = _get_val(values, layout, 'quartier')

            # Nettoyer le nom
            nom = clean_name(nom)
            if not is_valid_client_name(nom):
                stats['skipped'] += 1
                continue

            # Nettoyer le prix
            prix = _parse_price(prix_raw)

            # Nettoyer le forfait
            forfait = clean_name(forfait)

            # Si pas de forfait, essayer de l'inférer du prix
            if not forfait:
                if prix and prix > 0:
                    forfait = infer_plan_from_price(prix)
                else:
                    forfait = 'Bsic'
                    prix = 2000

            # Si pas de prix, l'inférer du forfait
            if prix is None or prix <= 0:
                normalized = normalize_plan_name(forfait)
                prix = get_default_price(normalized)

            # Créer/récupérer le quartier
            quartier_name = clean_name(quartier_name)
            if quartier_name:
                quartier = get_or_create_quartier(quartier_name, zone)
            else:
                # Utiliser le quartier par défaut (nom du fichier, ex: DIATA ZONE A)
                quartier = get_or_create_quartier(fallback_quartier, zone)

            # Créer/récupérer le plan
            plan = get_or_create_plan(forfait, prix)

            # Créer/récupérer le client
            rue_clean = clean_name(rue)
            client = get_or_create_client(nom, tel, tel2, quartier, zone, address=rue_clean)

            if client is None:
                stats['skipped'] += 1
                continue

            # Parser les dates
            date_debut = safe_parse_date(date_debut_raw) or date.today()
            date_fin = safe_parse_date(date_fin_raw)
            if not date_fin:
                date_fin = date_debut + timedelta(days=30)

            # Créer l'abonnement (get_or_create pour éviter les doublons)
            sub, sub_created = Subscription.objects.get_or_create(
                client=client,
                plan=plan,
                defaults={
                    'end_date': date_fin,
                    'is_active': True,
                }
            )
            if not sub_created:
                # Mettre à jour la date de fin si nécessaire
                if date_fin and sub.end_date != date_fin:
                    sub.end_date = date_fin
                    sub.save()
                stats['updated'] += 1
                print(f"    ♻️  {nom} | {plan.name} ({prix:.0f} FCFA) — déjà existant")
            else:
                stats['created'] += 1
                print(f"    ✅ {nom} | {plan.name} ({prix:.0f} FCFA) | {date_debut} → {date_fin}")

        except Exception as e:
            stats['errors'] += 1
            print(f"    ❌ Erreur ligne {row_idx}: {str(e)}")
            continue

    return stats


def _get_val(values, layout, key):
    """Extraire une valeur d'une ligne selon le layout."""
    idx = layout.get(key)
    if idx is not None and idx < len(values):
        return values[idx]
    return None


def _parse_price(value):
    """Parser un prix depuis différents formats."""
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return float(value) if value > 0 else None
    if isinstance(value, str):
        # Enlever les espaces, FCFA, etc.
        cleaned = value.strip().upper()
        cleaned = cleaned.replace('FCFA', '').replace('XAF', '').replace('F', '')
        cleaned = cleaned.replace(' ', '').replace('.', '').replace(',', '.')
        try:
            price = float(cleaned)
            return price if price > 0 else None
        except ValueError:
            return None
    return None


# ──────────────────────────────────────────────
# Main Import Function
# ──────────────────────────────────────────────
def import_all_zones(dry_run=False):
    """
    Importer toutes les zones depuis les fichiers Excel dans le dossier scripts/.
    """
    print("=" * 70)
    print("🚀 IMPORT UNIVERSEL - TOUTES LES ZONES")
    print("=" * 70)
    print(f"📁 Dossier: {SCRIPTS_DIR}")
    print()

    # Trouver tous les fichiers Excel
    excel_files = sorted(SCRIPTS_DIR.glob('*.xlsx'))

    if not excel_files:
        print("⚠️  Aucun fichier Excel trouvé dans le dossier scripts/")
        return

    print(f"📑 {len(excel_files)} fichier(s) Excel trouvé(s):")
    for f in excel_files:
        zone_name = derive_zone_name(f.name)
        default_q = derive_default_quartier(f.name)
        print(f"   • {f.name}  →  Zone: {zone_name}  |  Quartier par défaut: {default_q}")
    print()

    # Snapshot avant import
    print("📸 État de la base AVANT import:")
    print(f"   🏙️  Zones: {Zone.objects.count()}")
    print(f"   🏘️  Quartiers: {Quartier.objects.count()}")
    print(f"   📋 Plans: {SubscriptionPlan.objects.count()}")
    print(f"   👥 Clients: {User.objects.filter(role=User.Role.CLIENT).count()}")
    print(f"   📄 Abonnements: {Subscription.objects.count()}")
    print()

    if dry_run:
        print("⚠️  MODE DRY RUN - Aucune modification ne sera effectuée")
        return

    # Traiter chaque fichier
    grand_total = {'created': 0, 'errors': 0, 'skipped': 0, 'updated': 0}

    for excel_file in excel_files:
        zone_name = derive_zone_name(excel_file.name)
        default_quartier = derive_default_quartier(excel_file.name)
        print(f"\n{'─' * 70}")
        print(f"📂 Traitement de: {excel_file.name}")
        print(f"🏙️  Zone: {zone_name}  |  Quartier par défaut: {default_quartier}")
        print(f"{'─' * 70}")

        # Créer/récupérer la zone
        zone, zone_created = get_or_create_zone(zone_name)
        if zone_created:
            print(f"   ✨ Zone '{zone_name}' CRÉÉE")
        else:
            print(f"   ♻️  Zone '{zone_name}' existe déjà")

        # Charger le fichier Excel
        try:
            wb = openpyxl.load_workbook(str(excel_file), data_only=True)
            print(f"   📊 Feuilles: {wb.sheetnames}")
        except Exception as e:
            print(f"   ❌ Impossible d'ouvrir {excel_file.name}: {e}")
            continue

        # Traiter chaque feuille
        file_total = {'created': 0, 'errors': 0, 'skipped': 0, 'updated': 0}

        for sheet_name in wb.sheetnames:
            sheet = wb[sheet_name]
            print(f"\n   📄 Feuille: {sheet_name} ({sheet.max_row} lignes, {sheet.max_column} colonnes)")

            stats = process_sheet(sheet, zone, zone_name, sheet_name, default_quartier_name=default_quartier)

            for key in file_total:
                file_total[key] += stats.get(key, 0)

        # Résumé par fichier
        print(f"\n   📊 Résumé {excel_file.name}:")
        print(f"      ✅ Créés: {file_total['created']}")
        print(f"      ❌ Erreurs: {file_total['errors']}")
        print(f"      ⏭️  Ignorés: {file_total['skipped']}")

        for key in grand_total:
            grand_total[key] += file_total[key]

        wb.close()

    # ──────────────────────────────────────────
    # Résumé final
    # ──────────────────────────────────────────
    print(f"\n{'=' * 70}")
    print(f"📊 RÉSUMÉ FINAL DE L'IMPORT")
    print(f"{'=' * 70}")
    print(f"  ✅ Abonnements créés: {grand_total['created']}")
    print(f"  ❌ Erreurs: {grand_total['errors']}")
    print(f"  ⏭️  Lignes ignorées: {grand_total['skipped']}")

    print(f"\n📸 État de la base APRÈS import:")
    print(f"  🏙️  Zones: {Zone.objects.count()}")
    zones = Zone.objects.all()
    for z in zones:
        clients = User.objects.filter(zone=z, role=User.Role.CLIENT).count()
        quartiers = z.quartiers.count()
        print(f"      • {z.name}: {clients} clients, {quartiers} quartiers")

    print(f"  🏘️  Quartiers: {Quartier.objects.count()}")
    print(f"  📋 Plans d'abonnement: {SubscriptionPlan.objects.count()}")
    for plan in SubscriptionPlan.objects.all().order_by('price'):
        subs_count = Subscription.objects.filter(plan=plan).count()
        print(f"      • {plan.name}: {plan.price:.0f} FCFA ({subs_count} abonnements)")

    print(f"  👥 Clients: {User.objects.filter(role=User.Role.CLIENT).count()}")
    print(f"  📄 Abonnements: {Subscription.objects.count()}")

    print(f"\n✅ Import terminé avec succès !")
    print(f"{'=' * 70}")


# ──────────────────────────────────────────────
# Entry Point
# ──────────────────────────────────────────────
if __name__ == '__main__':
    import argparse
    parser = argparse.ArgumentParser(description="Import universel des zones depuis Excel")
    parser.add_argument('--dry-run', action='store_true', help="Mode simulation (pas de modification)")
    args = parser.parse_args()

    import_all_zones(dry_run=args.dry_run)
