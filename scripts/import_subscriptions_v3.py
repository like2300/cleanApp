#!/usr/bin/env python3
"""
Script pour importer les abonnements depuis DIATA ZONE B.xlsx
Avec détection de colonnes fixe basée sur l'analyse du fichier
"""

import os
import sys
import django
from datetime import datetime, date, timedelta
from dateutil.parser import parse as parse_date

# Setup Django
sys.path.append('/Users/omerlinks/Desktop/project/CLEAN')
os.environ.setdefault('DJANGO_SETTINGS_MODULE', 'core.settings')
django.setup()

import openpyxl
from business.models import Zone, Quartier, SubscriptionPlan, Subscription
from accounts.models import User
from django.db.models import Q

def parse_excel_date(value):
    """Convertir une valeur Excel en date Python"""
    if value is None:
        return None
    if isinstance(value, (datetime, date)):
        return value.date() if isinstance(value, datetime) else value
    if isinstance(value, str):
        try:
            date_str = value.strip().lower()
            if date_str.startswith('le '):
                date_str = date_str[3:]
            return parse_date(date_str).date()
        except:
            pass
    return None

def clean_phone(phone):
    """Nettoyer un numéro de téléphone"""
    if phone is None:
        return None
    phone_str = str(phone).strip()
    if phone_str in ['////', '///', '//', '/', 'None', '']:
        return None
    # Garder uniquement les chiffres
    cleaned = ''.join(c for c in phone_str if c.isdigit())
    return cleaned if cleaned and len(cleaned) >= 8 else None

def clean_name(name):
    """Nettoyer un nom"""
    if name is None:
        return None
    name_str = str(name).strip()
    if not name_str or name_str in ['None', '']:
        return None
    return ' '.join(name_str.split())

def get_or_create_zone(name):
    """Obtenir ou créer une zone"""
    try:
        return Zone.objects.get(name__iexact=name.strip())
    except Zone.DoesNotExist:
        return Zone.objects.create(name=name.strip())

def get_or_create_quartier(name, zone):
    """Obtenir ou créer un quartier"""
    name_clean = clean_name(name)
    if not name_clean:
        return None
    try:
        return Quartier.objects.get(name__iexact=name_clean, zone=zone)
    except Quartier.DoesNotExist:
        return Quartier.objects.create(name=name_clean, zone=zone)

def get_or_create_plan(name, price):
    """Obtenir ou créer un plan d'abonnement"""
    name_clean = clean_name(name)
    if not name_clean:
        name_clean = "Standard"
    
    # Normaliser les noms courants
    name_normalized = name_clean.strip()
    if 'basic' in name_normalized.lower():
        name_normalized = 'Bsic'
    elif 'standard' in name_normalized.lower():
        name_normalized = 'Standard'
    elif 'entreprise' in name_normalized.lower():
        name_normalized = 'Entreprise'
    elif '3 plus' in name_normalized.lower():
        name_normalized = '3 Plus'
    elif 'bsic' in name_normalized.lower():
        name_normalized = 'Bsic'
    
    try:
        return SubscriptionPlan.objects.get(name__iexact=name_normalized)
    except SubscriptionPlan.DoesNotExist:
        return SubscriptionPlan.objects.create(
            name=name_normalized,
            price=price,
            duration_days=30,
            description=f"Plan {name_normalized} - {price} FCFA"
        )

def get_or_create_client(name, phone, phone2, quartier, address):
    """Obtenir ou créer un client"""
    name_clean = clean_name(name)
    phone_clean = clean_phone(phone) if phone else None
    phone2_clean = clean_phone(phone2) if phone2 else None
    
    # Si pas de nom, utiliser le téléphone comme nom
    if not name_clean and phone_clean:
        name_clean = f"Client_{phone_clean[-4:]}"
    elif not name_clean:
        name_clean = "Client_Inconnu"
    
    # Vérifier si le client existe déjà par téléphone
    existing = User.objects.filter(
        Q(phone_number=phone_clean) | Q(phone_number_2=phone_clean),
        role=User.Role.CLIENT
    ).first()
    
    if existing:
        if phone2_clean and not existing.phone_number_2:
            existing.phone_number_2 = phone2_clean
        if quartier and not existing.quartier:
            existing.quartier = quartier
        if address and not existing.address:
            existing.address = address
        existing.save()
        return existing
    
    # Vérifier par nom d'utilisateur
    existing = User.objects.filter(
        username__iexact=name_clean,
        role=User.Role.CLIENT
    ).first()
    
    if existing:
        if phone_clean and not existing.phone_number:
            existing.phone_number = phone_clean
        if phone2_clean and not existing.phone_number_2:
            existing.phone_number_2 = phone2_clean
        if quartier and not existing.quartier:
            existing.quartier = quartier
        if address and not existing.address:
            existing.address = address
        existing.save()
        return existing
    
    # Créer un nouveau client
    return User.objects.create_user(
        username=name_clean,
        phone_number=phone_clean,
        phone_number_2=phone2_clean,
        quartier=quartier,
        address=address,
        role=User.Role.CLIENT
    )

def process_row(row_values, zone):
    """Traiter une ligne de données"""
    # Structure attendue pour Feuil1:
    # [N, numéro_rue, rue, quartier, nom, tel, tel2, forfait, prix, devise, date_debut, jour_recouv, date_recouv, ...]
    
    # Nettoyer les valeurs
    n = row_values[0] if len(row_values) > 0 else None  # N
    rue_num = row_values[1] if len(row_values) > 1 else None  # numéro de rue
    rue = clean_name(row_values[2]) if len(row_values) > 2 else None  # nom de la rue
    quartier_name = clean_name(row_values[3]) if len(row_values) > 3 else None  # quartier
    nom = clean_name(row_values[4]) if len(row_values) > 4 else None  # nom et prénom
    tel = clean_phone(row_values[5]) if len(row_values) > 5 else None  # téléphone
    tel2 = clean_phone(row_values[6]) if len(row_values) > 6 else None  # téléphone 2
    forfait = clean_name(row_values[7]) if len(row_values) > 7 else None  # forfait
    prix = row_values[8] if len(row_values) > 8 and isinstance(row_values[8], (int, float)) and row_values[8] > 0 else None  # prix
    devise = row_values[9] if len(row_values) > 9 else None  # devise (FCFA)
    date_debut_str = row_values[10] if len(row_values) > 10 else None  # date de début
    jour_recouv = row_values[11] if len(row_values) > 11 else None  # jour du recouvrement
    date_recouv_str = row_values[12] if len(row_values) > 12 else None  # date de recouvrement
    
    # Si prix n'est pas trouvé à la colonne 8, essayer d'autres colonnes
    if prix is None:
        for i in range(len(row_values)):
            val = row_values[i]
            if isinstance(val, (int, float)) and val > 0 and val < 100000:  # Prix raisonnable
                if i not in [0, 1, 11]:  # Exclure N, numéro_rue, jour_recouv
                    prix = float(val)
                    break
    
    # Si prix toujours None, utiliser 2000 par défaut
    if prix is None:
        prix = 2000
    
    # Si forfait est None ou vide, essayer de le déduire du prix
    if not forfait:
        if prix >= 10000:
            forfait = "Entreprise"
        elif prix >= 5000:
            forfait = "Standard"
        elif prix >= 2500:
            forfait = "Standard"
        else:
            forfait = "Bsic"
    
    # Si nom est None ou vide, sauter
    if not nom or nom.strip() in ['DIATA', 'DIATA ZONE B', 'Avenue 5fevrier', 'Rue massangui', 'Rue Boupanda', 'Rue Mbila', 'Rue Mont Fouary', 'Rue Makabana', 'rue Mbongui']:
        return None
    
    # Si quartier_name n'est pas DIATA, essayer de l'utiliser comme quartier
    if quartier_name and quartier_name.strip().lower() != 'diata':
        quartier_obj = get_or_create_quartier(quartier_name, zone)
    else:
        quartier_obj = get_or_create_quartier("DIATA", zone)
    
    # Si rue n'est pas None, l'utiliser comme adresse
    address = rue
    
    # Obtenir ou créer le plan
    plan = get_or_create_plan(forfait, prix)
    
    # Obtenir ou créer le client
    client = get_or_create_client(nom, tel, tel2, quartier_obj, address)
    
    # Parser les dates
    date_debut = parse_excel_date(date_debut_str) if date_debut_str else date.today()
    date_recouvrement = parse_excel_date(date_recouv_str) if date_recouv_str else None
    
    # Si date_recouvrement n'existe pas, utiliser date_debut + 30 jours
    if date_recouvrement is None:
        date_recouvrement = date_debut + timedelta(days=30)
    
    return {
        'client': client,
        'plan': plan,
        'prix': prix,
        'date_debut': date_debut,
        'date_fin': date_recouvrement,
        'nom': nom,
        'tel': tel
    }

def process_sheet_v2(sheet, zone_name="DIATA"):
    """Traiter une feuille Excel avec structure connue"""
    print(f"\nProcessing sheet: {sheet.title}")
    
    zone = get_or_create_zone(zone_name)
    print(f"Zone: {zone.name}")
    
    created = 0
    errors = 0
    skipped = 0
    
    for row_idx, row in enumerate(sheet.iter_rows(), 1):
        try:
            row_values = [cell.value for cell in row]
            
            # Sauter les lignes vides
            if all(v is None or (isinstance(v, str) and not v.strip()) for v in row_values):
                continue
            
            # Traiter la ligne
            data = process_row(row_values, zone)
            
            if data is None:
                skipped += 1
                continue
            
            # Créer l'abonnement
            subscription = Subscription.objects.create(
                client=data['client'],
                plan=data['plan'],
                start_date=data['date_debut'],
                end_date=data['date_fin'],
                is_active=True
            )
            
            created += 1
            print(f"  Created: {data['nom']} | {data['plan'].name} ({data['prix']} FCFA) | {data['date_debut']} to {data['date_fin']}")
            
        except Exception as e:
            errors += 1
            print(f"  Error row {row_idx}: {str(e)}")
            continue
    
    return created, errors, skipped

def process_sheet_feuil2(sheet, zone_name="DIATA"):
    """Traiter Feuil2 qui a une structure différente"""
    print(f"\nProcessing sheet: {sheet.title} (Feuil2 structure)")
    
    zone = get_or_create_zone(zone_name)
    print(f"Zone: {zone.name}")
    
    created = 0
    errors = 0
    skipped = 0
    
    for row_idx, row in enumerate(sheet.iter_rows(), 1):
        try:
            row_values = [cell.value for cell in row]
            
            # Sauter les lignes vides
            if all(v is None or (isinstance(v, str) and not v.strip()) for v in row_values):
                continue
            
            # Feuil2 semble commencer à la colonne C (index 2)
            # Structure: [None, None, N, Ruelle, Quartier, Nom, Tel, Tel2, Forfait, Prix, FCFA, Date Debut, ..., Date Recouv]
            
            if len(row_values) < 10:
                skipped += 1
                continue
            
            # Extraire à partir de la colonne 2
            n = row_values[2] if len(row_values) > 2 else None
            rue = clean_name(row_values[3]) if len(row_values) > 3 else None
            quartier_name = clean_name(row_values[4]) if len(row_values) > 4 else None
            nom = clean_name(row_values[5]) if len(row_values) > 5 else None
            tel = clean_phone(row_values[6]) if len(row_values) > 6 else None
            tel2 = clean_phone(row_values[7]) if len(row_values) > 7 else None
            forfait = clean_name(row_values[8]) if len(row_values) > 8 else None
            prix = row_values[9] if len(row_values) > 9 and isinstance(row_values[9], (int, float)) and row_values[9] > 0 else None
            devise = row_values[10] if len(row_values) > 10 else None
            date_debut_str = row_values[11] if len(row_values) > 11 else None
            date_recouv_str = row_values[13] if len(row_values) > 13 else None
            
            # Si prix est None, essayer de trouver une valeur numérique
            if prix is None:
                for i in range(len(row_values)):
                    val = row_values[i]
                    if isinstance(val, (int, float)) and val > 0 and val < 100000:
                        prix = float(val)
                        break
            
            if prix is None:
                prix = 2000
            
            if not forfait:
                if prix >= 10000:
                    forfait = "Entreprise"
                elif prix >= 5000:
                    forfait = "Standard"
                else:
                    forfait = "Bsic"
            
            # Sauter les lignes qui ne sont pas des données clients
            if not nom or nom.strip() in ['DIATA', 'DIATA ZONE B', 'Avenue 5fevrier', 'Rue massangui', 'Rue Boupanda', 'Rue Mbila', 'Rue Mont Fouary', 'Rue Makabana', 'rue Mbongui']:
                skipped += 1
                continue
            
            # Obtenir ou créer le quartier
            if quartier_name and quartier_name.strip().lower() != 'diata':
                quartier_obj = get_or_create_quartier(quartier_name, zone)
            else:
                quartier_obj = get_or_create_quartier("DIATA", zone)
            
            # Obtenir ou créer le plan
            plan = get_or_create_plan(forfait, prix)
            
            # Obtenir ou créer le client
            client = get_or_create_client(nom, tel, tel2, quartier_obj, rue)
            
            # Parser les dates
            date_debut = parse_excel_date(date_debut_str) if date_debut_str else date.today()
            date_recouvrement = parse_excel_date(date_recouv_str) if date_recouv_str else None
            
            if date_recouvrement is None:
                date_recouvrement = date_debut + timedelta(days=30)
            
            # Créer l'abonnement
            subscription = Subscription.objects.create(
                client=client,
                plan=plan,
                start_date=date_debut,
                end_date=date_recouvrement,
                is_active=True
            )
            
            created += 1
            print(f"  Created: {nom} | {plan.name} ({prix} FCFA) | {date_debut} to {date_recouvrement}")
            
        except Exception as e:
            errors += 1
            print(f"  Error row {row_idx}: {str(e)}")
            continue
    
    return created, errors, skipped

def main():
    """Fonction principale"""
    excel_path = '/Users/omerlinks/Desktop/project/CLEAN/scripts/DIATA ZONE B.xlsx'
    
    if not os.path.exists(excel_path):
        print(f"Error: File not found: {excel_path}")
        sys.exit(1)
    
    print(f"Loading Excel file: {excel_path}")
    
    try:
        wb = openpyxl.load_workbook(excel_path)
        print(f"Sheets: {wb.sheetnames}")
        
        total_created = 0
        total_errors = 0
        total_skipped = 0
        
        for sheet_name in wb.sheetnames:
            sheet = wb[sheet_name]
            if sheet_name == 'Feuil2':
                created, errors, skipped = process_sheet_feuil2(sheet, zone_name="DIATA")
            else:
                created, errors, skipped = process_sheet_v2(sheet, zone_name="DIATA")
            total_created += created
            total_errors += errors
            total_skipped += skipped
        
        print(f"\n=== Summary ===")
        print(f"Subscriptions created: {total_created}")
        print(f"Errors: {total_errors}")
        print(f"Skipped rows: {total_skipped}")
        
        print(f"\n=== Database Status ===")
        print(f"Zones: {Zone.objects.count()}")
        print(f"Quartiers: {Quartier.objects.count()}")
        print(f"Subscription Plans: {SubscriptionPlan.objects.count()}")
        print(f"Clients: {User.objects.filter(role=User.Role.CLIENT).count()}")
        print(f"Subscriptions: {Subscription.objects.count()}")
        
        # Afficher les plans créés
        print(f"\nPlans d'abonnement:")
        for plan in SubscriptionPlan.objects.all():
            subs_count = Subscription.objects.filter(plan=plan).count()
            print(f"  - {plan.name}: {plan.price} FCFA ({subs_count} abonnements)")
        
        # Afficher quelques clients
        print(f"\nPremiers clients créés:")
        clients = User.objects.filter(role=User.Role.CLIENT).order_by('-id')[:10]
        for client in clients:
            subs = Subscription.objects.filter(client=client).count()
            print(f"  - {client.username} ({client.phone_number}) - {subs} abonnement(s)")
        
    except Exception as e:
        print(f"Error: {str(e)}")
        import traceback
        traceback.print_exc()
        sys.exit(1)

if __name__ == '__main__':
    main()
