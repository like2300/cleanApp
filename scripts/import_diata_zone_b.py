#!/usr/bin/env python3
"""
Script final pour importer correctement les abonnements depuis DIATA ZONE B.xlsx
"""

import os
import sys
import django
from datetime import datetime, date, timedelta
from dateutil.parser import parse as parse_date

# Setup Django
sys.path.append('/Users/omerlinks/Desktop/project/CLEAN')
os.environ.setdefault('DJANGO_SETTINGS_MODULE', 'core.settings')
django.setup()

import openpyxl
from business.models import Zone, Quartier, SubscriptionPlan, Subscription
from accounts.models import User
from django.db.models import Q

def parse_date(value):
    """Convertir une valeur en date"""
    if value is None:
        return None
    if isinstance(value, (datetime, date)):
        return value.date() if isinstance(value, datetime) else value
    if isinstance(value, str):
        try:
            date_str = value.strip().lower()
            if date_str.startswith('le '):
                date_str = date_str[3:]
            return parse_date(date_str).date()
        except:
            pass
    return None

def clean_phone(phone):
    """Nettoyer un numéro de téléphone"""
    if phone is None:
        return None
    phone_str = str(phone).strip()
    if phone_str in ['////', '///', '//', '/', 'None', '']:
        return None
    cleaned = ''.join(c for c in phone_str if c.isdigit())
    return cleaned if cleaned and len(cleaned) >= 8 else None

def clean_name(name):
    """Nettoyer un nom"""
    if name is None:
        return None
    name_str = str(name).strip()
    if not name_str:
        return None
    return ' '.join(name_str.split())

def get_or_create_zone(name):
    try:
        return Zone.objects.get(name__iexact=name.strip())
    except Zone.DoesNotExist:
        return Zone.objects.create(name=name.strip())

def get_or_create_quartier(name, zone):
    name_clean = clean_name(name)
    if not name_clean:
        return None
    try:
        return Quartier.objects.get(name__iexact=name_clean, zone=zone)
    except Quartier.DoesNotExist:
        return Quartier.objects.create(name=name_clean, zone=zone)

def get_or_create_plan(name, price):
    name_clean = clean_name(name)
    if not name_clean:
        name_clean = "Standard"
    
    # Normaliser
    name_lower = name_clean.lower()
    if 'basic' in name_lower or 'bsic' in name_lower:
        normalized = 'Bsic'
    elif 'standard' in name_lower:
        normalized = 'Standard'
    elif 'entreprise' in name_lower:
        normalized = 'Entreprise'
    elif '3 plus' in name_lower or 'plus' in name_lower:
        normalized = '3 Plus'
    else:
        normalized = name_clean
    
    try:
        return SubscriptionPlan.objects.get(name__iexact=normalized)
    except SubscriptionPlan.DoesNotExist:
        return SubscriptionPlan.objects.create(
            name=normalized,
            price=price,
            duration_days=30,
            description=f"Plan {normalized} - {price} FCFA"
        )

def get_or_create_client(name, phone, phone2, quartier, address):
    name_clean = clean_name(name)
    phone_clean = clean_phone(phone) if phone else None
    phone2_clean = clean_phone(phone2) if phone2 else None
    
    if not name_clean:
        if phone_clean:
            name_clean = f"Client_{phone_clean[-4:]}"
        else:
            name_clean = "Client_Inconnu"
    
    # Vérifier par téléphone
    existing = User.objects.filter(
        Q(phone_number=phone_clean) | Q(phone_number_2=phone_clean),
        role=User.Role.CLIENT
    ).first()
    
    if existing:
        if phone2_clean and not existing.phone_number_2:
            existing.phone_number_2 = phone2_clean
        if quartier and not existing.quartier:
            existing.quartier = quartier
        if address and not existing.address:
            existing.address = address
        existing.save()
        return existing
    
    # Vérifier par nom
    existing = User.objects.filter(
        username__iexact=name_clean,
        role=User.Role.CLIENT
    ).first()
    
    if existing:
        if phone_clean and not existing.phone_number:
            existing.phone_number = phone_clean
        if phone2_clean and not existing.phone_number_2:
            existing.phone_number_2 = phone2_clean
        if quartier and not existing.quartier:
            existing.quartier = quartier
        if address and not existing.address:
            existing.address = address
        existing.save()
        return existing
    
    return User.objects.create_user(
        username=name_clean,
        phone_number=phone_clean,
        phone_number_2=phone2_clean,
        quartier=quartier,
        address=address,
        role=User.Role.CLIENT
    )

def extract_data_from_row(values, sheet_type='feuil1'):
    """Extraire les données d'une ligne selon le type de feuille"""
    
    # Extraire tous les numéros de téléphone
    phones = []
    for val in values:
        if val and isinstance(val, str):
            cleaned = clean_phone(val)
            if cleaned:
                phones.append(cleaned)
    
    tel = phones[0] if phones else None
    tel2 = phones[1] if len(phones) > 1 else None
    
    # Extraire le nom - chercher la première chaîne qui n'est pas une adresse connue
    known_addresses = ['diata', 'avenue', 'av ', 'rue ', 'massangui', 'boupanda', 
                      'mbila', 'mont fouary', 'makabana', 'mbongui', 'makola', 'yamba',
                      'fevrier', 'chateau', 'libis', 'apoen', 'baconbgo']
    
    name = None
    for val in values:
        if val and isinstance(val, str):
            val_lower = val.strip().lower()
            if not any(addr in val_lower for addr in known_addresses):
                if not any(c.isdigit() for c in val.strip()):  # Pas de chiffres
                    name = clean_name(val)
                    break
    
    # Si pas de nom trouvé, prendre la première chaîne non vide
    if not name:
        for val in values:
            if val and isinstance(val, str) and val.strip():
                name = clean_name(val)
                break
    
    # Extraire le prix - chercher une valeur numérique > 0
    # Exclure les indices 0, 1, 2 (N, numéro de rue, etc.)
    prix = None
    for i, val in enumerate(values):
        if isinstance(val, (int, float)) and val > 0:
            # Exclure les petits nombres (1, 2, 3...) qui sont probablement des numéros de ligne
            if val > 100 or (val > 0 and i >= 8):
                prix = float(val)
                break
    
    # Si prix toujours None, chercher n'importe quelle valeur numérique
    if prix is None:
        for val in values:
            if isinstance(val, (int, float)) and val > 0:
                prix = float(val)
                break
    
    if prix is None:
        prix = 2000  # Default
    
    # Extraire le forfait
    forfait = None
    for val in values:
        if val and isinstance(val, str):
            val_lower = val.strip().lower()
            for kw in ['bsic', 'basic', 'standard', 'entreprise', '3 plus', 'plus', 'forfait', 'pack']:
                if kw in val_lower:
                    forfait = clean_name(val)
                    break
            if forfait:
                break
    
    # Si pas de forfait, déduire du prix
    if not forfait:
        if prix >= 10000:
            forfait = "Entreprise"
        elif prix >= 5000:
            forfait = "Standard"
        elif prix >= 2500:
            forfait = "Standard"
        else:
            forfait = "Bsic"
    
    # Extraire les dates
    dates = []
    for val in values:
        parsed = parse_date(val)
        if parsed:
            dates.append(parsed)
    
    date_debut = dates[0] if dates else date.today()
    date_fin = dates[1] if len(dates) > 1 else None
    
    # Si date_fin est None, utiliser date_debut + 30 jours
    if date_fin is None:
        date_fin = date_debut + timedelta(days=30)
    
    # Extraire la rue
    rue = None
    for val in values:
        if val and isinstance(val, str):
            val_lower = val.strip().lower()
            if any(addr in val_lower for addr in ['av ', 'rue ', 'avenue', 'massangui', 'boupanda', 
                      'mbila', 'mont fouary', 'makabana', 'mbongui', 'makola', 'yamba']):
                rue = clean_name(val)
                break
    
    # Déterminer le quartier
    quartier_name = "DIATA"
    for val in values:
        if val and isinstance(val, str) and 'DIATA' in val:
            quartier_name = "DIATA"
            break
    
    return {
        'name': name,
        'tel': tel,
        'tel2': tel2,
        'prix': prix,
        'forfait': forfait,
        'date_debut': date_debut,
        'date_fin': date_fin,
        'rue': rue,
        'quartier_name': quartier_name
    }

def process_sheet(sheet, zone):
    """Traiter une feuille"""
    created = 0
    errors = 0
    skipped = 0
    
    for row_idx, row in enumerate(sheet.iter_rows(), 1):
        try:
            values = [cell.value for cell in row]
            
            # Sauter les lignes vides
            if all(v is None or (isinstance(v, str) and not v.strip()) for v in values):
                continue
            
            # Sauter les lignes d'en-tête
            header_keywords = ['n', 'ruelle', 'quartier', 'nom', 'tel', 'forfait', 'pack', 
                             'prix', 'date', 'dte', 'num', 'colonne']
            if any(isinstance(v, str) and v.strip().lower() in header_keywords for v in values):
                continue
            
            # Sauter les lignes avec "DIATA ZONE B"
            if any(isinstance(v, str) and 'diata zone b' in v.lower() for v in values):
                continue
            
            # Extraire les données
            data = extract_data_from_row(values)
            
            # Sauter si pas de nom ou nom invalide
            if not data['name'] or len(data['name']) < 2:
                skipped += 1
                continue
            
            # Sauter les noms qui sont des adresses
            invalid_names = ['DIATA', 'Avenue 5fevrier', 'Rue massangui', 'Rue Boupanda', 
                           'Rue Mbila', 'Rue Mont Fouary', 'Rue Makabana', 'rue Mbongui',
                           'massangui', 'boupanda', 'mbila', 'mont fouary', 'makabana', 'mbongui']
            if any(data['name'].strip().lower() == name.lower() for name in invalid_names):
                skipped += 1
                continue
            
            # Obtenir ou créer le quartier
            quartier = get_or_create_quartier(data['quartier_name'], zone)
            
            # Obtenir ou créer le plan
            plan = get_or_create_plan(data['forfait'], data['prix'])
            
            # Obtenir ou créer le client
            client = get_or_create_client(data['name'], data['tel'], data['tel2'], quartier, data['rue'])
            
            # Créer l'abonnement
            subscription = Subscription.objects.create(
                client=client,
                plan=plan,
                start_date=data['date_debut'],
                end_date=data['date_fin'],
                is_active=True
            )
            
            created += 1
            print(f"  Created: {data['name']} | {plan.name} ({data['prix']} FCFA) | {data['date_debut']} to {data['date_fin']}")
            
        except Exception as e:
            errors += 1
            print(f"  Error row {row_idx}: {str(e)}")
            import traceback
            traceback.print_exc()
            continue
    
    return created, errors, skipped

def main():
    excel_path = '/Users/omerlinks/Desktop/project/CLEAN/scripts/DIATA ZONE B.xlsx'
    
    if not os.path.exists(excel_path):
        print(f"Error: File not found: {excel_path}")
        sys.exit(1)
    
    print(f"Loading Excel file: {excel_path}")
    
    try:
        wb = openpyxl.load_workbook(excel_path)
        print(f"Sheets: {wb.sheetnames}")
        
        zone = get_or_create_zone("DIATA")
        print(f"Using zone: {zone.name}")
        
        total_created = 0
        total_errors = 0
        total_skipped = 0
        
        for sheet_name in wb.sheetnames:
            sheet = wb[sheet_name]
            print(f"\nProcessing {sheet_name}...")
            created, errors, skipped = process_sheet(sheet, zone)
            total_created += created
            total_errors += errors
            total_skipped += skipped
        
        print(f"\n{'='*60}")
        print(f"SUMMARY")
        print(f"{'='*60}")
        print(f"Subscriptions created: {total_created}")
        print(f"Errors: {total_errors}")
        print(f"Skipped rows: {total_skipped}")
        
        print(f"\nDatabase Status:")
        print(f"  Zones: {Zone.objects.count()}")
        print(f"  Quartiers: {Quartier.objects.count()}")
        print(f"  Subscription Plans: {SubscriptionPlan.objects.count()}")
        print(f"  Clients: {User.objects.filter(role=User.Role.CLIENT).count()}")
        print(f"  Subscriptions: {Subscription.objects.count()}")
        
        # Afficher les plans
        print(f"\nPlans d'abonnement:")
        for plan in SubscriptionPlan.objects.all().order_by('price'):
            subs_count = Subscription.objects.filter(plan=plan).count()
            print(f"  - {plan.name}: {plan.price:.0f} FCFA ({subs_count} abonnements)")
        
        # Stats par forfait
        print(f"\nRépartition par forfait:")
        for plan in SubscriptionPlan.objects.all().order_by('name'):
            subs = Subscription.objects.filter(plan=plan)
            if subs.count() > 0:
                print(f"  {plan.name}: {subs.count()} abonnements")
        
        print(f"\n✓ Import terminé avec succès !")
        
    except Exception as e:
        print(f"Error: {str(e)}")
        import traceback
        traceback.print_exc()
        sys.exit(1)

if __name__ == '__main__':
    main()
