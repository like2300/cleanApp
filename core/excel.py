from decimal import Decimal

import openpyxl
from django.http import HttpResponse
from django.utils import timezone
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

from accounts.models import CompanySettings


class ExcelReportBuilder:
    """Reusable modern Excel report builder with consistent UI/UX."""

    CONTENT_TYPE = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

    def __init__(self, title, request=None, filename_prefix="rapport"):
        self.title = title
        self.request = request
        self.filename_prefix = filename_prefix
        self.company = CompanySettings.get_settings()
        self.wb = openpyxl.Workbook()
        self._table_index = 1

        self.primary = self._clean_color(self.company.primary_color, "010694")
        self.secondary = self._clean_color(self.company.secondary_color, "2435C9")
        self.light_fill = "F8FAFC"
        self.border_color = "E2E8F0"
        self.text_color = "0F172A"
        self.muted_color = "64748B"

        self.thin_border = Border(
            left=Side(style="thin", color=self.border_color),
            right=Side(style="thin", color=self.border_color),
            top=Side(style="thin", color=self.border_color),
            bottom=Side(style="thin", color=self.border_color),
        )

    def _clean_color(self, value, fallback):
        value = (value or fallback).replace("#", "").upper()
        return value if len(value) == 6 else fallback

    def _safe_sheet_title(self, title):
        for char in ["\\", "/", "?", "*", "[", "]", ":"]:
            title = title.replace(char, "-")
        return title[:31] or "Rapport"

    def _safe_value(self, value):
        if value is None:
            return "-"
        if isinstance(value, Decimal):
            return float(value)
        return value

    def _safe_table_ref(self, ws, start_row, start_col, end_row, end_col):
        return f"{get_column_letter(start_col)}{start_row}:{get_column_letter(end_col)}{end_row}"

    def active_sheet(self, title):
        ws = self.wb.active
        ws.title = self._safe_sheet_title(title)
        return ws

    def create_sheet(self, title):
        return self.wb.create_sheet(title=self._safe_sheet_title(title))

    def add_title(self, ws, columns_count, subtitle=None):
        columns_count = max(columns_count, 2)
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=columns_count)
        title_cell = ws.cell(row=1, column=1, value=self.title)
        title_cell.font = Font(bold=True, size=16, color="FFFFFF")
        title_cell.fill = PatternFill(
            start_color=self.primary, end_color=self.primary, fill_type="solid"
        )
        title_cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 30

        generated_by = "-"
        if self.request and getattr(self.request, "user", None):
            generated_by = getattr(self.request.user, "username", "-")

        meta = (
            subtitle
            or f"{self.company.name} • Généré le {timezone.now().strftime('%d/%m/%Y %H:%M')} • Par {generated_by}"
        )
        ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=columns_count)
        meta_cell = ws.cell(row=2, column=1, value=meta)
        meta_cell.font = Font(size=10, italic=True, color=self.muted_color)
        meta_cell.fill = PatternFill(
            start_color=self.light_fill, end_color=self.light_fill, fill_type="solid"
        )
        meta_cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[2].height = 22

    def add_filters_summary(self, ws, filters, start_row=4):
        ws.cell(row=start_row, column=1, value="Filtres appliqués").font = Font(
            bold=True, color=self.text_color
        )
        row = start_row + 1
        clean_filters = [
            (label, value) for label, value in filters if value not in [None, "", []]
        ]
        if not clean_filters:
            clean_filters = [("Aucun filtre", "Toutes les données autorisées")]
        for label, value in clean_filters:
            ws.cell(row=row, column=1, value=label).font = Font(
                bold=True, color=self.muted_color
            )
            ws.cell(row=row, column=2, value=str(value))
            row += 1
        return row + 1

    def add_kpis(self, ws, kpis, start_row=4, start_col=1):
        row = start_row
        col = start_col
        for index, (label, value) in enumerate(kpis):
            current_col = col + (index * 2)
            ws.cell(row=row, column=current_col, value=label).font = Font(
                bold=True, size=9, color="FFFFFF"
            )
            ws.cell(row=row, column=current_col).fill = PatternFill(
                start_color=self.secondary, end_color=self.secondary, fill_type="solid"
            )
            value_cell = ws.cell(
                row=row + 1, column=current_col, value=self._safe_value(value)
            )
            value_cell.font = Font(bold=True, size=13, color=self.text_color)
            value_cell.fill = PatternFill(
                start_color=self.light_fill,
                end_color=self.light_fill,
                fill_type="solid",
            )
            ws.column_dimensions[get_column_letter(current_col)].width = 18
        return row + 3

    def add_table(self, ws, columns, rows, start_row=5, start_col=1, table_name=None):
        header_fill = PatternFill(
            start_color=self.primary, end_color=self.primary, fill_type="solid"
        )
        header_font = Font(color="FFFFFF", bold=True)
        center = Alignment(horizontal="center", vertical="center")

        for col_offset, column in enumerate(columns):
            cell = ws.cell(row=start_row, column=start_col + col_offset, value=column)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center
            cell.border = self.thin_border

        data_start = start_row + 1
        row_count = 0
        for row_offset, row in enumerate(rows):
            row_count += 1
            for col_offset, value in enumerate(row):
                cell = ws.cell(
                    row=data_start + row_offset,
                    column=start_col + col_offset,
                    value=self._safe_value(value),
                )
                cell.border = self.thin_border
                cell.alignment = Alignment(vertical="top", wrap_text=True)
                if row_offset % 2 == 1:
                    cell.fill = PatternFill(
                        start_color=self.light_fill,
                        end_color=self.light_fill,
                        fill_type="solid",
                    )
                if isinstance(value, Decimal):
                    cell.number_format = "#,##0"

        end_row = max(start_row + row_count, start_row + 1)
        end_col = start_col + len(columns) - 1
        if row_count == 0:
            ws.cell(row=data_start, column=start_col, value="Aucune donnée")

        ref = self._safe_table_ref(ws, start_row, start_col, end_row, end_col)
        table = Table(displayName=table_name or f"Tableau{self._table_index}", ref=ref)
        self._table_index += 1
        style = TableStyleInfo(
            name="TableStyleMedium2",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False,
        )
        table.tableStyleInfo = style
        try:
            ws.add_table(table)
        except ValueError:
            pass

        ws.freeze_panes = ws.cell(row=start_row + 1, column=start_col).coordinate
        ws.auto_filter.ref = ref
        self.auto_size(ws)
        return end_row + 2

    def add_total_row(self, ws, row, label_col, label, values):
        ws.cell(row=row, column=label_col, value=label).font = Font(
            bold=True, color=self.text_color
        )
        for col, value in values.items():
            cell = ws.cell(row=row, column=col, value=self._safe_value(value))
            cell.font = Font(bold=True, color=self.text_color)
            cell.fill = PatternFill(
                start_color="E0F2FE", end_color="E0F2FE", fill_type="solid"
            )
            cell.border = self.thin_border
            if isinstance(value, (int, float, Decimal)):
                cell.number_format = "#,##0"
        self.auto_size(ws)

    def auto_size(self, ws, max_width=45):
        for column_cells in ws.columns:
            column_letter = get_column_letter(column_cells[0].column)
            max_length = 0
            for cell in column_cells:
                if cell.value is not None:
                    max_length = max(max_length, len(str(cell.value)))
            ws.column_dimensions[column_letter].width = min(
                max(max_length + 2, 12), max_width
            )

    def finalize(self):
        for ws in self.wb.worksheets:
            ws.sheet_view.showGridLines = False
            self.auto_size(ws)

    def response(self):
        self.finalize()
        response = HttpResponse(content_type=self.CONTENT_TYPE)
        timestamp = timezone.now().strftime("%Y%m%d_%H%M")
        response["Content-Disposition"] = (
            f'attachment; filename="{self.filename_prefix}_{timestamp}.xlsx"'
        )
        self.wb.save(response)
        return response
