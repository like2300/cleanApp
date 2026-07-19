from calendar import monthrange
from datetime import timedelta

from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.db.models import Q
from django.shortcuts import get_object_or_404, redirect, render
from django.utils import timezone

from accounts.models import User
from finance.billing import (
    invoice_due_date_for_subscription,
    next_period_end_date,
    payment_subscription_end_date,
)
from finance.models import Invoice, Payment

from .models import Employee, Position, Subscription, SubscriptionPlan, Zone


def add_months_safe(date_value, months):
    """Return date_value shifted by N months while keeping a valid day."""
    month_index = date_value.month - 1 + months
    year = date_value.year + month_index // 12
    month = month_index % 12 + 1
    day = min(date_value.day, monthrange(year, month)[1])
    return date_value.replace(year=year, month=month, day=day)


def calculate_next_available_invoice_due_date(
    client, subscription, reference_date=None
):
    """Find the next due month without an existing invoice for this client/subscription."""
    reference_date = reference_date or timezone.now().date()
    due_date = Invoice.calculate_due_date(client, reference_date)

    while Invoice.objects.filter(
        client=client,
        subscription=subscription,
        due_date__year=due_date.year,
        due_date__month=due_date.month,
    ).exists():
        due_date = add_months_safe(due_date, 1)

    return due_date


def _create_renewal_invoice_if_missing(subscription):
    """Create the pending renewal invoice for an expired subscription once.

    Returns None when the subscription has no plan (e.g. the plan was retired
    via SET_NULL) — without a price there is nothing to invoice.
    """
    if subscription.plan_id is None:
        return None

    if Invoice.objects.filter(
        client=subscription.client,
        subscription=subscription,
        status=Invoice.Status.PENDING,
        due_date=invoice_due_date_for_subscription(subscription),
    ).exists():
        return None

    return Invoice.objects.create(
        client=subscription.client,
        subscription=subscription,
        amount=subscription.plan.price,
        invoice_type=Invoice.InvoiceType.PAIEMENT,
        due_date=invoice_due_date_for_subscription(subscription),
        status=Invoice.Status.PENDING,
        synced=False,
    )


# Add this utility function to check and deactivate expired subscriptions
def check_and_deactivate_expired_subscriptions():
    """Check expired subscriptions, create renewal invoices, then deactivate them."""
    today = timezone.now().date()

    expired_subscriptions = Subscription.objects.filter(
        is_active=True,
        end_date__lte=today,
    )

    for subscription in expired_subscriptions:
        _create_renewal_invoice_if_missing(subscription)
        subscription.is_active = False
        subscription.synced = False
        subscription.save(update_fields=["is_active", "synced", "updated_at"])

    return expired_subscriptions.count()


def check_and_deactivate_expired_momo_subscriptions():
    """MoMo uses the same saved monthly due date as all other clients."""
    return 0


from datetime import timedelta
from decimal import Decimal

import openpyxl
from django.db.models import Q
from django.http import HttpResponse, JsonResponse
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from accounts.models import CompanySettings
from core.excel import ExcelReportBuilder
from core.utils import (
    actionnaire_read_only,
    block_read_only_role,
    check_zone_access,
    get_zone_queryset,
)


@login_required
def zone_subscriptions_calendar(request, pk):
    """Calendar view of subscriptions showing payment status by month"""
    if not check_zone_access(request.user, pk):
        messages.error(request, "Accès refusé.")
        return redirect("dashboard")

    zone = get_object_or_404(Zone, pk=pk)

    # Get filter parameters
    quartier_filter = request.GET.get("quartier")
    period_filter = request.GET.get("period", "current")
    search_query = request.GET.get("search")

    # Get clients with their subscriptions
    clients = User.objects.filter(zone=zone, role=User.Role.CLIENT)

    # Apply filters
    if quartier_filter:
        clients = clients.filter(quartier_id=quartier_filter)

    if search_query:
        clients = clients.filter(
            Q(first_name__icontains=search_query)
            | Q(username__icontains=search_query)
            | Q(registration_number__icontains=search_query)
        )

    # Generate months based on period filter
    if period_filter == "current":
        # Current month + next 5 months
        months = [
            timezone.now().replace(day=1),
            (timezone.now().replace(day=1) + timedelta(days=32)).replace(day=1),
            (timezone.now().replace(day=1) + timedelta(days=62)).replace(day=1),
            (timezone.now().replace(day=1) + timedelta(days=93)).replace(day=1),
            (timezone.now().replace(day=1) + timedelta(days=124)).replace(day=1),
            (timezone.now().replace(day=1) + timedelta(days=155)).replace(day=1),
        ]
    else:
        # All months for the year
        months = []
        for m in range(1, 13):
            months.append(timezone.datetime(timezone.now().year, m, 1))

    context = {
        "zone": zone,
        "clients": clients,
        "months": months,
        "quartiers": zone.quartiers.all(),
        "selected_quartier": quartier_filter,
        "selected_period": period_filter,
        "search_query": search_query,
    }

    return render(request, "business/zone_subscriptions_calendar.html", context)


@login_required
def export_zone_subscriptions_excel(request, pk):
    """Export subscriptions to Excel with filtering."""
    if not check_zone_access(request.user, pk):
        messages.error(request, "Accès refusé.")
        return redirect("dashboard")

    zone = get_object_or_404(Zone, pk=pk)
    quartier_filter = request.GET.get("quartier")
    status_filter = request.GET.get("status")
    search_query = request.GET.get("search")

    subscriptions = Subscription.objects.filter(client__zone=zone).select_related(
        "client", "client__quartier", "plan"
    )
    if quartier_filter:
        subscriptions = subscriptions.filter(client__quartier_id=quartier_filter)
    if status_filter == "active":
        subscriptions = subscriptions.filter(is_active=True)
    elif status_filter == "inactive":
        subscriptions = subscriptions.filter(is_active=False)
    if search_query:
        subscriptions = subscriptions.filter(
            Q(client__first_name__icontains=search_query)
            | Q(client__username__icontains=search_query)
            | Q(client__registration_number__icontains=search_query)
        )

    report = ExcelReportBuilder(
        f"Abonnements - {zone.name}",
        request=request,
        filename_prefix=f"abonnements_{zone.name}",
    )
    ws = report.active_sheet("Abonnements")
    report.add_title(ws, 10)
    start = report.add_filters_summary(
        ws,
        [
            ("Zone", zone.name),
            ("Quartier", quartier_filter),
            ("Statut", status_filter),
            ("Recherche", search_query),
        ],
    )
    start = report.add_kpis(
        ws,
        [
            ("Abonnements", subscriptions.count()),
            ("Actifs", subscriptions.filter(is_active=True).count()),
        ],
        start,
    )
    rows = [
        [
            sub.client.id,
            sub.client.first_name or sub.client.username,
            sub.client.registration_number or sub.client.username,
            sub.client.phone_number or "-",
            sub.client.email or "-",
            sub.client.quartier.name if sub.client.quartier else "-",
            sub.plan.name,
            sub.plan.price,
            sub.end_date.strftime("%d/%m/%Y") if sub.end_date else "-",
            "Actif" if sub.is_active else "Inactif",
        ]
        for sub in subscriptions
    ]
    report.add_table(
        ws,
        [
            "ID",
            "Nom",
            "Matricule",
            "Téléphone",
            "Email",
            "Quartier",
            "Plan",
            "Prix",
            "Échéance",
            "Statut",
        ],
        rows,
        start_row=start,
        table_name="AbonnementsZone",
    )
    return report.response()


# Import the template tag function for payment checking
def has_payment_for_month(client, month):
    """Check if client has payment for specific month"""
    from business.templatetags.has_payment import (
        has_payment_for_month as original_has_payment_for_month,
    )

    return original_has_payment_for_month(client, month)


@login_required
def zone_subscriptions_manage(request, pk):
    """Manage all subscriptions for a zone with filtering and export"""
    if not check_zone_access(request.user, pk):
        messages.error(request, "Accès refusé à cette zone.")
        return redirect("dashboard")

    zone = get_object_or_404(Zone, pk=pk)

    # Get filter parameters
    quartier_filter = request.GET.get("quartier")
    status_filter = request.GET.get("status")
    search_query = request.GET.get("search")
    period_filter = request.GET.get("period", "current")  # Default to current month
    year_filter = request.GET.get("year")
    payment_status_filter = request.GET.get("payment_status")

    # Base queryset - get all clients in the zone
    clients = User.objects.filter(zone=zone, role=User.Role.CLIENT)

    # Apply filters
    if quartier_filter:
        clients = clients.filter(quartier_id=quartier_filter)

    if status_filter:
        if status_filter == "active":
            clients = clients.filter(is_active=True)
        elif status_filter == "inactive":
            clients = clients.filter(is_active=False)

    if search_query:
        clients = clients.filter(
            Q(first_name__icontains=search_query)
            | Q(username__icontains=search_query)
            | Q(registration_number__icontains=search_query)
        )

    # Apply payment status filters after getting all clients
    if payment_status_filter:
        # Generate months if not already defined (for payment status filtering)
        if "months" not in locals():
            current_year = timezone.now().year
            target_year = int(year_filter) if year_filter else current_year

            if period_filter == "current":
                # Current month + next 5 months
                months = []
                for i in range(6):
                    month_date = (
                        timezone.now().replace(day=1) + timedelta(days=31 * i)
                    ).replace(day=1)
                    # If year filter is specified, use that year
                    if year_filter:
                        month_date = month_date.replace(year=target_year)
                    months.append(month_date)
            else:
                # All months for the selected year
                months = []
                for m in range(1, 13):
                    months.append(timezone.datetime(target_year, m, 1))

        filtered_clients = []
        for client in clients:
            if payment_status_filter == "all_paid":
                # Check if client has paid all months in the selected period
                all_months_paid = True
                for month in months:
                    if not has_payment_for_month(client, month):
                        all_months_paid = False
                        break
                if all_months_paid:
                    filtered_clients.append(client)
            elif payment_status_filter == "irregular":
                # Check if client has irregular payments (some paid, some not)
                has_some_paid = False
                has_some_unpaid = False
                for month in months:
                    if has_payment_for_month(client, month):
                        has_some_paid = True
                    else:
                        has_some_unpaid = True
                    if has_some_paid and has_some_unpaid:
                        break
                if has_some_paid and has_some_unpaid:
                    filtered_clients.append(client)
        clients = filtered_clients

    # Get related data for display (only if clients is still a QuerySet)
    if hasattr(clients, "select_related"):
        clients = clients.select_related("quartier").prefetch_related(
            "subscriptions", "validated_payments"
        )

    # Ensure months is defined for the template context (it might have been defined in payment status filtering)
    if "months" not in locals():
        current_year = timezone.now().year
        target_year = int(year_filter) if year_filter else current_year

    if period_filter == "current":
        # Current month + next 5 months
        months = []
        for i in range(6):
            month_date = (
                timezone.now().replace(day=1) + timedelta(days=31 * i)
            ).replace(day=1)
            # If year filter is specified, use that year
            if year_filter:
                month_date = month_date.replace(year=target_year)
            months.append(month_date)
    else:
        # All months for the selected year
        months = []
        for m in range(1, 13):
            months.append(timezone.datetime(target_year, m, 1))

    context = {
        "zone": zone,
        "clients": clients,
        "months": months,
        "quartiers": zone.quartiers.all(),
        "selected_quartier": quartier_filter,
        "selected_status": status_filter,
        "selected_period": period_filter,
        "selected_year": year_filter,
        "selected_payment_status": payment_status_filter,
        "search_query": search_query,
    }

    return render(request, "business/zone_subscriptions_manage.html", context)


@login_required
def employee_list(request):
    if request.user.role == User.Role.ZONE_MANAGER:
        messages.error(request, "Accès refusé aux ressources globales du personnel.")
        return redirect("dashboard")

    employees = get_zone_queryset(
        request.user, Employee.objects.all().select_related("position", "zone")
    )

    # Get available roles based on user permissions
    if request.user.role == User.Role.ZONE_MANAGER:
        available_roles = [
            (User.Role.AGENT, "Agent")
        ]  # Zone managers can only create AGENTS
    else:
        available_roles = [
            (User.Role.SUPER_ADMIN, "Super Admin"),
            (User.Role.ZONE_MANAGER, "Chef de Zone"),
            (User.Role.ACCOUNTANT, "Comptable"),
            (User.Role.AGENT, "Agent"),
            (User.Role.CLIENT, "Client"),
            (User.Role.SHAREHOLDER, "Actionnaire"),
        ]

    zones = get_zone_queryset(request.user, Zone.objects.all())
    return render(
        request,
        "business/employee_list.html",
        {"employees": employees, "roles": available_roles, "zones": zones},
    )


from django.http import HttpResponse, JsonResponse


@login_required
def export_employees_excel(request):
    if request.user.role == User.Role.ZONE_MANAGER:
        messages.error(request, "Accès refusé.")
        return redirect("dashboard")

    # Get filters from GET parameters
    search = request.GET.get("search", "").lower()
    zone_id = request.GET.get("zone", "")
    position_id = request.GET.get("position", "")

    employees = get_zone_queryset(
        request.user, Employee.objects.all().select_related("position", "zone")
    )

    if zone_id:
        employees = employees.filter(zone_id=zone_id)
    if position_id:
        employees = employees.filter(position_id=position_id)

    # Filter by search term in Python to match the JS behavior or use Q objects
    if search:
        from django.db.models import Q

        employees = employees.filter(
            Q(first_name__icontains=search)
            | Q(last_name__icontains=search)
            | Q(phone_number__icontains=search)
            | Q(position__title__icontains=search)
            | Q(zone__name__icontains=search)
        )

    report = ExcelReportBuilder(
        "Rapport des employés", request=request, filename_prefix="employes"
    )
    ws = report.active_sheet("Employés")
    report.add_title(ws, 7)
    start = report.add_filters_summary(
        ws,
        [
            ("Recherche", search),
            ("Zone", zone_id),
            ("Poste", position_id),
        ],
    )
    total_salary = sum(emp.salary for emp in employees)
    start = report.add_kpis(
        ws,
        [("Effectif", employees.count()), ("Masse salariale", total_salary)],
        start,
    )
    rows = [
        [
            emp.last_name,
            emp.first_name,
            emp.phone_number or "-",
            emp.position.title if emp.position else "-",
            emp.zone.name if emp.zone else "-",
            emp.salary,
            emp.hired_at.strftime("%d/%m/%Y") if emp.hired_at else "-",
        ]
        for emp in employees
    ]
    report.add_table(
        ws,
        ["Nom", "Prénom", "Téléphone", "Poste", "Zone", "Salaire", "Date embauche"],
        rows,
        start_row=start,
        table_name="Employes",
    )
    return report.response()


from django.http import JsonResponse


@login_required
def employee_create(request):
    if request.user.role == User.Role.ZONE_MANAGER:
        messages.error(request, "Accès refusé.")
        return redirect("dashboard")

    if request.method == "POST":
        first_name = request.POST.get("first_name")
        last_name = request.POST.get("last_name")
        phone = request.POST.get("phone")
        phone_2 = request.POST.get("phone_2")
        address = request.POST.get("address")
        zone_id = request.POST.get("zone")
        role = request.POST.get("role")  # Changed from position to role

        # Restriction: only SUPER_ADMIN and ACCOUNTANT can set salary
        if request.user.role in [User.Role.SUPER_ADMIN, User.Role.ACCOUNTANT]:
            salary = request.POST.get("salary", 0)
        else:
            salary = 0

        hired_at = request.POST.get("hired_at")
        photo = request.FILES.get("photo")

        zone = get_object_or_404(Zone, id=zone_id)

        # Create or get position based on role
        position, created = Position.objects.get_or_create(title=role.upper())

        Employee.objects.create(
            first_name=first_name,
            last_name=last_name,
            photo=photo,
            phone_number=phone,
            phone_number_2=phone_2,
            address=address,
            zone=zone,
            salary=salary,
            position=position,
            hired_at=hired_at,
            synced=False,
        )
        messages.success(request, "Employé ajouté avec succès.")
        return redirect("employee_list")

    zones = Zone.objects.all()
    # Get available roles based on user permissions
    if request.user.role == User.Role.ZONE_MANAGER:
        available_roles = [
            (User.Role.AGENT, "Agent")
        ]  # Zone managers can only create AGENTS
    else:
        available_roles = [
            (User.Role.SUPER_ADMIN, "Super Admin"),
            (User.Role.ZONE_MANAGER, "Chef de Zone"),
            (User.Role.ACCOUNTANT, "Comptable"),
            (User.Role.AGENT, "Agent"),
            (User.Role.CLIENT, "Client"),
            (User.Role.SHAREHOLDER, "Actionnaire"),
        ]

    return render(
        request,
        "business/employee_form.html",
        {"zones": zones, "roles": available_roles},
    )


from django.http import JsonResponse


@login_required
def employee_edit(request, pk):
    employee = get_object_or_404(Employee, pk=pk)
    if employee.zone and not check_zone_access(request.user, employee.zone.id):
        messages.error(request, "Accès refusé.")
        return redirect("dashboard")
    if request.method == "POST":
        employee.first_name = request.POST.get("first_name")
        employee.last_name = request.POST.get("last_name")
        employee.phone_number = request.POST.get("phone")
        employee.phone_number_2 = request.POST.get("phone_2")
        employee.address = request.POST.get("address")
        zone_id = request.POST.get("zone")

        # Restriction: only SUPER_ADMIN and ACCOUNTANT can modify salary
        if request.user.role in [User.Role.SUPER_ADMIN, User.Role.ACCOUNTANT]:
            employee.salary = request.POST.get("salary", employee.salary)

        position_id = request.POST.get("position")
        employee.hired_at = request.POST.get("hired_at")

        if request.FILES.get("photo"):
            employee.photo = request.FILES.get("photo")

        employee.zone = get_object_or_404(Zone, id=zone_id)
        employee.position = get_object_or_404(Position, id=position_id)
        employee.synced = False
        employee.save()

        messages.success(request, f"Employé {employee.first_name} mis à jour.")
        if request.user.role == User.Role.ZONE_MANAGER:
            return redirect("zone_employees_manage", pk=employee.zone.pk)
        return redirect("employee_list")

    zones = Zone.objects.all()
    positions = Position.objects.all()
    return render(
        request,
        "business/employee_form.html",
        {"employee": employee, "zones": zones, "positions": positions, "is_edit": True},
    )


from django.http import JsonResponse


@login_required
def employee_print(request, pk):
    employee = get_object_or_404(Employee, pk=pk)
    if employee.zone and not check_zone_access(request.user, employee.zone.id):
        messages.error(request, "Accès refusé.")
        return redirect("dashboard")
    return render(
        request,
        "business/employee_print.html",
        {"employee": employee, "now": timezone.now()},
    )


from django.http import JsonResponse


@login_required
def employee_delete(request, pk):
    employee = get_object_or_404(Employee, pk=pk)
    if employee.zone and not check_zone_access(request.user, employee.zone.id):
        messages.error(request, "Accès refusé.")
        return redirect("dashboard")
    if request.method == "POST":
        zone_pk = employee.zone.pk if employee.zone else None
        employee.delete()
        messages.success(request, "Employé supprimé.")
        if request.user.role == User.Role.ZONE_MANAGER and zone_pk:
            return redirect("zone_employees_manage", pk=zone_pk)
        return redirect("employee_list")
    return render(
        request, "business/employee_confirm_delete.html", {"employee": employee}
    )


from django.http import JsonResponse


@login_required
def employee_detail(request, pk):
    employee = get_object_or_404(Employee, pk=pk)
    if employee.zone and not check_zone_access(request.user, employee.zone.id):
        messages.error(request, "Accès refusé.")
        return redirect("dashboard")
    return render(
        request,
        "business/employee_detail.html",
        {"employee": employee, "now": timezone.now()},
    )


from django.http import JsonResponse


@login_required
def position_list(request):
    positions = Position.objects.all()
    return render(request, "business/position_list.html", {"positions": positions})


from django.http import JsonResponse


@login_required
def position_create(request):
    if request.method == "POST":
        title = request.POST.get("title")
        if title:
            Position.objects.create(title=title)
            messages.success(request, f"Poste '{title}' ajouté.")
        else:
            messages.error(request, "Le titre du poste est requis.")
    return redirect(request.META.get("HTTP_REFERER", "employee_list"))


from django.http import JsonResponse


@login_required
def position_edit(request, pk):
    position = get_object_or_404(Position, pk=pk)
    if request.method == "POST":
        position.title = request.POST.get("title")
        position.save()
        messages.success(request, f"Poste '{position.title}' mis à jour.")
        return redirect("position_list")
    return render(
        request, "business/position_form.html", {"position": position, "is_edit": True}
    )


from django.http import JsonResponse


@login_required
def position_delete(request, pk):
    position = get_object_or_404(Position, pk=pk)
    if request.method == "POST":
        position.delete()
        messages.success(request, "Poste supprimé.")
        return redirect("position_list")
    return render(
        request, "business/position_confirm_delete.html", {"position": position}
    )


from django.http import JsonResponse


@login_required
def subscription_plan_list(request):
    if request.user.role not in [User.Role.SUPER_ADMIN, User.Role.ACCOUNTANT]:
        messages.error(request, "Accès refusé.")
        return redirect("dashboard")
    plans = SubscriptionPlan.objects.all()
    return render(request, "business/subscription_plan_list.html", {"plans": plans})


from django.http import JsonResponse


@login_required
def subscription_plan_create(request):
    if request.user.role != User.Role.SUPER_ADMIN:
        messages.error(
            request, "Accès refusé. Seul le super administrateur peut créer des plans."
        )
        return redirect("subscription_plan_list")
    if request.method == "POST":
        name = request.POST.get("name")
        price = request.POST.get("price")
        duration = request.POST.get("duration_days")
        description = request.POST.get("description", "")
        characteristics = request.POST.get("characteristics", "")

        SubscriptionPlan.objects.create(
            name=name,
            price=price,
            duration_days=duration,
            description=description,
            characteristics=characteristics,
        )
        messages.success(request, f"Plan '{name}' créé.")
        return redirect(request.META.get("HTTP_REFERER", "subscription_plan_list"))
    return redirect("subscription_plan_list")


from django.http import JsonResponse


@login_required
def subscription_plan_edit(request, pk):
    if request.user.role not in [User.Role.SUPER_ADMIN, User.Role.ACCOUNTANT]:
        messages.error(request, "Accès refusé.")
        return redirect("dashboard")
    plan = get_object_or_404(SubscriptionPlan, pk=pk)
    if request.method == "POST":
        plan.name = request.POST.get("name")
        plan.price = request.POST.get("price")
        plan.duration_days = request.POST.get("duration_days")
        plan.description = request.POST.get("description")
        plan.characteristics = request.POST.get("characteristics")
        plan.save()
        messages.success(request, f"Plan '{plan.name}' mis à jour.")
        return redirect("subscription_plan_list")
    return render(
        request, "business/subscription_plan_form.html", {"plan": plan, "is_edit": True}
    )


from django.http import JsonResponse


@login_required
def subscription_plan_delete(request, pk):
    if request.user.role not in [User.Role.SUPER_ADMIN, User.Role.ACCOUNTANT]:
        messages.error(request, "Accès refusé.")
        return redirect("dashboard")
    plan = get_object_or_404(SubscriptionPlan, pk=pk)
    if request.method == "POST":
        # Don't nuke the billing history of clients still subscribed; the
        # Subscription.plan FK is on_delete=SET_NULL, so the rows are kept
        # with plan_id=None after the plan is gone.
        active_subs = plan.subscription_set.filter(is_active=True)
        if active_subs.exists():
            messages.error(
                request,
                f"Impossible de supprimer ce plan : {active_subs.count()} abonnement(s) "
                "actif(s) y est/sont rattaché(s). Clôturez d’abord ces abonnements.",
            )
            return redirect("subscription_plan_list")
        plan.delete()
        messages.success(request, "Plan supprimé.")
        return redirect("subscription_plan_list")
    return render(
        request, "business/subscription_plan_confirm_delete.html", {"plan": plan}
    )


from django.http import JsonResponse


@login_required
def subscription_list(request):
    subscriptions = get_zone_queryset(
        request.user, Subscription.objects.all(), zone_field="client__zone"
    )
    return render(
        request, "business/subscription_list.html", {"subscriptions": subscriptions}
    )


@login_required
def export_subscriptions_excel(request):
    """Export all subscriptions to Excel with filtering."""
    subscriptions = get_zone_queryset(
        request.user, Subscription.objects.all(), zone_field="client__zone"
    ).select_related("client", "client__quartier", "plan")

    status_filter = request.GET.get("status")
    search_query = request.GET.get("search")

    if status_filter == "active":
        subscriptions = subscriptions.filter(is_active=True)
    elif status_filter == "inactive":
        subscriptions = subscriptions.filter(is_active=False)

    if search_query:
        subscriptions = subscriptions.filter(
            Q(client__first_name__icontains=search_query)
            | Q(client__username__icontains=search_query)
            | Q(client__registration_number__icontains=search_query)
        )

    report = ExcelReportBuilder(
        "Tous les Abonnements",
        request=request,
        filename_prefix="tous_abonnements",
    )
    ws = report.active_sheet("Abonnements")
    report.add_title(ws, 10)
    start = report.add_filters_summary(
        ws,
        [
            ("Statut", status_filter),
            ("Recherche", search_query),
        ],
    )
    start = report.add_kpis(
        ws,
        [
            ("Abonnements", subscriptions.count()),
            ("Actifs", subscriptions.filter(is_active=True).count()),
        ],
        start,
    )
    rows = [
        [
            sub.client.id,
            sub.client.first_name or sub.client.username,
            sub.client.registration_number or sub.client.username,
            sub.client.phone_number or "-",
            sub.client.email or "-",
            sub.client.quartier.name if sub.client.quartier else "-",
            sub.client.zone.name if sub.client.zone else "-",
            sub.plan.name,
            sub.plan.price,
            sub.end_date.strftime("%d/%m/%Y") if sub.end_date else "-",
            "Actif" if sub.is_active else "Inactif",
        ]
        for sub in subscriptions
    ]
    report.add_table(
        ws,
        [
            "ID",
            "Nom",
            "Matricule",
            "Téléphone",
            "Email",
            "Quartier",
            "Zone",
            "Plan",
            "Prix",
            "Échéance",
            "Statut",
        ],
        rows,
        start_row=start,
        table_name="TousAbonnements",
    )
    return report.response()


from django.db.models import Count, Q
from django.http import JsonResponse


@login_required
def zone_list(request):
    zones = get_zone_queryset(
        request.user,
        Zone.objects.annotate(
            employee_count=Count("employees", distinct=True),
            client_count=Count(
                "user", filter=Q(user__role=User.Role.CLIENT), distinct=True
            ),
        ),
    ).order_by("name")
    return render(request, "business/zone_list.html", {"zones": zones})


from django.http import JsonResponse


@login_required
def zone_create(request):
    # Only Super Admin should create zones
    if request.user.role != User.Role.SUPER_ADMIN:
        messages.error(request, "Accès refusé.")
        return redirect("zone_list")

    if request.method == "POST":
        name = request.POST.get("name")
        description = request.POST.get("description", "")
        if name:
            Zone.objects.create(name=name, description=description)
            messages.success(request, f"Zone '{name}' ajoutée.")
        else:
            messages.error(request, "Le nom de la zone est requis.")
    return redirect("zone_list")


@login_required
def zone_edit(request, pk):
    # Only Super Admin can edit zones
    if request.user.role != User.Role.SUPER_ADMIN:
        messages.error(request, "Accès refusé.")
        return redirect("zone_list")

    zone = get_object_or_404(Zone, pk=pk)

    if request.method == "POST":
        name = request.POST.get("name")
        description = request.POST.get("description", "")
        if name:
            zone.name = name
            zone.description = description
            zone.save()
            messages.success(request, f"Zone '{name}' modifiée avec succès.")
            return redirect("zone_list")
        else:
            messages.error(request, "Le nom de la zone est requis.")

    return render(request, "business/zone_edit.html", {"zone": zone})


@login_required
def zone_delete(request, pk):
    # Only Super Admin can delete zones
    if request.user.role != User.Role.SUPER_ADMIN:
        messages.error(request, "Accès refusé.")
        return redirect("zone_list")

    zone = get_object_or_404(Zone, pk=pk)

    if request.method == "POST":
        zone_name = zone.name
        zone.delete()
        messages.success(request, f"Zone '{zone_name}' supprimée avec succès.")
        return redirect("zone_list")

    return render(request, "business/zone_confirm_delete.html", {"zone": zone})


from django.http import JsonResponse


@login_required
def zone_detail(request, pk):
    # Check if user has access to this zone
    if not check_zone_access(request.user, pk):
        messages.error(request, "Accès refusé à cette zone.")
        return redirect("dashboard")

    zone = get_object_or_404(Zone, pk=pk)
    employees = zone.employees.all()
    clients = User.objects.filter(zone=zone, role=User.Role.CLIENT)
    invoices = Invoice.objects.filter(client__zone=zone).order_by("-created_at")

    positions = Position.objects.all()
    subscription_plans = SubscriptionPlan.objects.all()
    subscriptions = Subscription.objects.filter(client__zone=zone).order_by(
        "-start_date"
    )

    # Financial Summary
    total_invoiced = sum(inv.amount for inv in invoices)
    total_paid = sum(inv.amount for inv in invoices if inv.status == "PAID")

    # Performance Calculation
    payment_ratio = (total_paid / total_invoiced * 100) if total_invoiced > 0 else 0

    # Global average for comparison
    all_invoices = Invoice.objects.all()
    global_invoiced = sum(inv.amount for inv in all_invoices)
    global_paid = sum(inv.amount for inv in all_invoices if inv.status == "PAID")
    global_payment_ratio = (
        (global_paid / global_invoiced * 100) if global_invoiced > 0 else 0
    )

    is_above_average = payment_ratio >= global_payment_ratio

    context = {
        "zone": zone,
        "employees": employees,
        "clients": clients,
        "invoices": invoices,
        "positions": positions,
        "subscription_plans": subscription_plans,
        "subscriptions": subscriptions,
        "total_invoiced": total_invoiced,
        "total_paid": total_paid,
        "payment_ratio": round(payment_ratio, 1),
        "is_above_average": is_above_average,
        "global_payment_ratio": round(global_payment_ratio, 1),
    }

    return render(request, "business/zone_detail.html", context)


@login_required
def clients_without_zone(request):
    """View to display and assign clients without a zone to Diata B or other zones"""
    if request.user.role != User.Role.SUPER_ADMIN:
        messages.error(request, "Accès refusé. Réservé aux Super Admin.")
        return redirect("dashboard")

    # Get all clients without a zone
    clients_without_zone = (
        User.objects.filter(role=User.Role.CLIENT, zone__isnull=True)
        .select_related("quartier")
        .order_by("username")
    )

    # Get all zones for the dropdown
    zones = Zone.objects.all().order_by("name")

    if request.method == "POST":
        zone_id = request.POST.get("zone_id")
        client_ids = request.POST.getlist("client_ids")

        if zone_id and client_ids:
            try:
                zone = get_object_or_404(Zone, pk=zone_id)
                selected_clients = User.objects.filter(
                    id__in=client_ids, role=User.Role.CLIENT
                )
                count = selected_clients.update(zone=zone)
                messages.success(
                    request,
                    f"{count} clients assignés à la zone {zone.name} avec succès.",
                )
            except Exception as e:
                messages.error(request, f"Erreur: {e}")
        return redirect("clients_without_zone")

    return render(
        request,
        "business/clients_without_zone.html",
        {
            "clients": clients_without_zone,
            "zones": zones,
        },
    )


from django.http import JsonResponse


@login_required
def zone_clients_manage(request, pk):
    # Check if user has access to this zone
    if not check_zone_access(request.user, pk):
        messages.error(request, "Accès refusé à cette zone.")
        return redirect("dashboard")

    # Check for expired subscriptions
    expired_count = check_and_deactivate_expired_subscriptions()
    expired_momo_count = check_and_deactivate_expired_momo_subscriptions()

    total_expired = expired_count + expired_momo_count
    if total_expired > 0:
        messages.info(
            request,
            f"{total_expired} abonnement(s) expiré(s) ont été désactivés automatiquement.",
        )

    zone = get_object_or_404(Zone, pk=pk)
    clients = User.objects.filter(zone=zone, role=User.Role.CLIENT).prefetch_related(
        "invoices", "subscriptions", "subscriptions__plan"
    )

    client_list = []
    for client in clients:
        pending_invoice = client.invoices.filter(status=Invoice.Status.PENDING).first()
        client_list.append({"client": client, "pending_invoice": pending_invoice})

    subscription_plans = SubscriptionPlan.objects.all()
    quartiers = zone.quartiers.all()
    return render(
        request,
        "business/zone_clients_manage.html",
        {
            "zone": zone,
            "client_list": client_list,
            "subscription_plans": subscription_plans,
            "quartiers": quartiers,
        },
    )


from django.http import JsonResponse


@login_required
def export_zone_clients_excel(request, pk):
    if not check_zone_access(request.user, pk):
        messages.error(request, "Accès refusé.")
        return redirect("dashboard")
    zone = get_object_or_404(Zone, pk=pk)
    search = request.GET.get("search", "").lower()

    clients = User.objects.filter(zone=zone, role=User.Role.CLIENT).prefetch_related(
        "subscriptions", "invoices"
    )

    if search:
        from django.db.models import Q

        clients = clients.filter(
            Q(username__icontains=search)
            | Q(email__icontains=search)
            | Q(phone_number__icontains=search)
            | Q(registration_number__icontains=search)
        )

    # Get Company Colors
    settings = CompanySettings.get_settings()
    primary_color = settings.primary_color.replace("#", "")
    secondary_color = settings.secondary_color.replace("#", "")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Liste des Clients"

    # Styles
    header_fill = PatternFill(
        start_color=primary_color, end_color=primary_color, fill_type="solid"
    )
    header_font = Font(color="FFFFFF", bold=True, size=12)
    sub_header_fill = PatternFill(
        start_color=secondary_color, end_color=secondary_color, fill_type="solid"
    )
    center_align = Alignment(horizontal="center", vertical="center")
    border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    # Summary Section
    ws.merge_cells("A1:I1")
    ws["A1"] = f"RAPPORT D'ANALYSE CLIENTS - ZONE: {zone.name.upper()}"
    ws["A1"].font = Font(bold=True, size=16, color=primary_color)
    ws["A1"].alignment = center_align

    ws["A3"] = "Date du rapport:"
    ws["B3"] = timezone.now().strftime("%d/%m/%Y %H:%M")
    ws["A4"] = "Total Clients:"
    ws["B4"] = clients.count()

    # Table Header
    columns = [
        "Matricule",
        "Client",
        "Email",
        "Téléphone",
        "Plan Actuel",
        "Durée (Jours)",
        "Stabilité (%)",
        "Régularité",
        "Statut",
    ]

    header_row = 6
    for col_num, column_title in enumerate(columns, 1):
        cell = ws.cell(row=header_row, column=col_num)
        cell.value = column_title
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_align
        cell.border = border

    # Data Rows
    for row_num, client in enumerate(clients, header_row + 1):
        # Analytics Calculation
        sub = client.subscriptions.filter(is_active=True).first()
        total_inv = client.invoices.count()
        paid_inv = client.invoices.filter(status="PAID").count()

        stability = (paid_inv / total_inv * 100) if total_inv > 0 else 0
        duration = (timezone.now().date() - client.date_joined.date()).days

        # Regularity Logic
        if sub and stability >= 80:
            regularity = "Très Régulier"
            reg_color = "00B050"  # Green
        elif sub or stability >= 50:
            regularity = "Moyen"
            reg_color = "FFC000"  # Orange
        else:
            regularity = "Irrégulier"
            reg_color = "FF0000"  # Red

        data = [
            client.registration_number or "-",
            client.username,
            client.email or "-",
            client.phone_number or "-",
            sub.plan.name if sub else "Aucun",
            duration,
            f"{round(stability, 1)}%",
            regularity,
            "Actif" if sub else "Inactif",
        ]

        for col_num, cell_value in enumerate(data, 1):
            cell = ws.cell(row=row_num, column=col_num)
            cell.value = cell_value
            cell.border = border
            cell.alignment = Alignment(vertical="center")

            # Highlight Regularity column
            if col_num == 8:
                cell.font = Font(color=reg_color, bold=True)

            # Highlight Status
            if col_num == 9:
                if cell_value == "Actif":
                    cell.font = Font(color="00B050", bold=True)
                else:
                    cell.font = Font(color="FF0000", bold=True)

    # Adjust Column Widths
    for col in ws.columns:
        max_length = 0
        # Get the first cell of the column that is NOT a MergedCell if possible,
        # or just use the column index
        col_letter = get_column_letter(col[0].column)

        for cell in col:
            # Skip cells that are part of a merge but not the master cell (MergedCell)
            # as they usually have None value or would interfere
            if hasattr(cell, "value") and cell.value:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass

        adjusted_width = max_length + 2
        if adjusted_width > 50:
            adjusted_width = 50  # Cap width
        ws.column_dimensions[col_letter].width = adjusted_width

    # Second Sheet for Detailed Analysis
    ws2 = wb.create_sheet(title="Analyse de Performance")
    ws2["A1"] = "INDICATEURS DE PERFORMANCE DE LA ZONE"
    ws2["A1"].font = Font(bold=True, size=14)

    # Calculate global zone metrics
    total_clients = clients.count()
    active_clients = clients.filter(subscriptions__is_active=True).distinct().count()
    total_invoiced = sum(
        inv.amount for client in clients for inv in client.invoices.all()
    )
    total_paid = sum(
        inv.amount
        for client in clients
        for inv in client.invoices.filter(status="PAID")
    )
    zone_stability = (total_paid / total_invoiced * 100) if total_invoiced > 0 else 0

    analysis_data = [
        ["Indicateur", "Valeur"],
        [
            "Taux d'Activité",
            f"{round(active_clients / total_clients * 100, 1)}%"
            if total_clients > 0
            else "0%",
        ],
        ["Stabilité Financière Zone", f"{round(zone_stability, 1)}%"],
        ["Nombre de Clients Actifs", active_clients],
        ["Chiffre d'Affaires Encaissé", f"{total_paid} FCFA"],
    ]

    for r_num, row_data in enumerate(analysis_data, 3):
        for c_num, val in enumerate(row_data, 1):
            c = ws2.cell(row=r_num, column=c_num)
            c.value = val
            c.border = border
            if r_num == 3:
                c.fill = sub_header_fill
                c.font = Font(color="FFFFFF", bold=True)

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = (
        f"attachment; filename=clients_{zone.name}_{timezone.now().strftime('%Y%m%d')}.xlsx"
    )
    wb.save(response)
    return response


from django.http import JsonResponse


@login_required
def export_zone_employees_excel(request, pk):
    if not check_zone_access(request.user, pk):
        messages.error(request, "Accès refusé.")
        return redirect("dashboard")
    zone = get_object_or_404(Zone, pk=pk)
    search = request.GET.get("search", "").lower()

    employees = zone.employees.all().select_related("position")

    if search:
        from django.db.models import Q

        employees = employees.filter(
            Q(first_name__icontains=search)
            | Q(last_name__icontains=search)
            | Q(phone_number__icontains=search)
            | Q(position__title__icontains=search)
        )

    # Get Company Colors
    settings = CompanySettings.get_settings()
    primary_color = settings.primary_color.replace("#", "")
    secondary_color = settings.secondary_color.replace("#", "")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Liste du Personnel"

    # Styles
    header_fill = PatternFill(
        start_color=primary_color, end_color=primary_color, fill_type="solid"
    )
    header_font = Font(color="FFFFFF", bold=True, size=12)
    center_align = Alignment(horizontal="center", vertical="center")
    border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    # Summary Section
    ws.merge_cells("A1:G1")
    ws["A1"] = f"RAPPORT DU PERSONNEL - ZONE: {zone.name.upper()}"
    ws["A1"].font = Font(bold=True, size=16, color=primary_color)
    ws["A1"].alignment = center_align

    ws["A3"] = "Date du rapport:"
    ws["B3"] = timezone.now().strftime("%d/%m/%Y %H:%M")
    ws["A4"] = "Effectif Total:"
    ws["B4"] = employees.count()

    # Table Header
    columns = [
        "Nom",
        "Prénom",
        "Téléphone",
        "Poste",
        "Salaire (FCFA)",
        "Ancienneté (Mois)",
        "Date Embauche",
    ]

    header_row = 6
    for col_num, column_title in enumerate(columns, 1):
        cell = ws.cell(row=header_row, column=col_num)
        cell.value = column_title
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_align
        cell.border = border

    # Data Rows
    for row_num, emp in enumerate(employees, header_row + 1):
        # Calculation: Seniority in months
        today = timezone.now().date()
        months_seniority = (today.year - emp.hired_at.year) * 12 + (
            today.month - emp.hired_at.month
        )

        data = [
            emp.last_name.upper(),
            emp.first_name,
            emp.phone_number or "-",
            emp.position.title if emp.position else "-",
            emp.salary,
            months_seniority,
            emp.hired_at.strftime("%d/%m/%Y"),
        ]

        for col_num, cell_value in enumerate(data, 1):
            cell = ws.cell(row=row_num, column=col_num)
            cell.value = cell_value
            cell.border = border
            cell.alignment = Alignment(vertical="center")

            # Format Salary
            if col_num == 5:
                cell.number_format = "#,##0"

    # Adjust Column Widths
    for col in ws.columns:
        max_length = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if hasattr(cell, "value") and cell.value:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
        adjusted_width = max_length + 2
        if adjusted_width > 40:
            adjusted_width = 40
        ws.column_dimensions[col_letter].width = adjusted_width

    # Second Sheet: Analysis
    ws2 = wb.create_sheet(title="Analyse RH")
    ws2["A1"] = "ANALYSE DE LA MASSE SALARIALE - " + zone.name.upper()
    ws2["A1"].font = Font(bold=True, size=14)

    total_salary = sum(emp.salary for emp in employees)
    avg_salary = (total_salary / employees.count()) if employees.count() > 0 else 0

    analysis_data = [
        ["Indicateur", "Valeur"],
        ["Masse Salariale Mensuelle", f"{total_salary} FCFA"],
        ["Salaire Moyen", f"{round(avg_salary, 0)} FCFA"],
        ["Effectif", employees.count()],
    ]

    for r_num, row_data in enumerate(analysis_data, 3):
        for c_num, val in enumerate(row_data, 1):
            c = ws2.cell(row=r_num, column=c_num)
            c.value = val
            c.border = border
            if r_num == 3:
                c.fill = PatternFill(
                    start_color=secondary_color,
                    end_color=secondary_color,
                    fill_type="solid",
                )
                c.font = Font(color="FFFFFF", bold=True)

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = (
        f"attachment; filename=employees_{zone.name}_{timezone.now().strftime('%Y%m%d')}.xlsx"
    )
    wb.save(response)
    return response


from django.http import JsonResponse


@login_required
@actionnaire_read_only
def zone_employees_manage(request, pk):
    if request.user.role == User.Role.ZONE_MANAGER:
        messages.error(request, "Accès refusé aux ressources du personnel.")
        return redirect("dashboard")

    if not check_zone_access(request.user, pk):
        messages.error(request, "Accès refusé.")
        return redirect("dashboard")
    zone = get_object_or_404(Zone, pk=pk)
    employees = zone.employees.all().select_related("position")

    # Get available roles based on user permissions
    if request.user.role == User.Role.ZONE_MANAGER:
        available_roles = [
            (User.Role.AGENT, "Agent")
        ]  # Zone managers can only create AGENTS
    else:
        available_roles = [
            (User.Role.SUPER_ADMIN, "Super Admin"),
            (User.Role.ZONE_MANAGER, "Chef de Zone"),
            (User.Role.ACCOUNTANT, "Comptable"),
            (User.Role.AGENT, "Agent"),
            (User.Role.CLIENT, "Client"),
            (User.Role.SHAREHOLDER, "Actionnaire"),
        ]

    return render(
        request,
        "business/zone_employees_manage.html",
        {"zone": zone, "employees": employees, "roles": available_roles},
    )


from django.http import JsonResponse


@login_required
def zone_finance_manage(request, pk):
    if not check_zone_access(request.user, pk):
        messages.error(request, "Accès refusé.")
        return redirect("dashboard")
    zone = get_object_or_404(Zone, pk=pk)
    clients = User.objects.filter(zone=zone, role=User.Role.CLIENT)
    invoices = Invoice.objects.filter(client__zone=zone).order_by("-created_at")
    subscriptions = Subscription.objects.filter(
        client__zone=zone, is_active=True
    ).select_related("plan", "client")
    subscription_plans = SubscriptionPlan.objects.all()

    # Financial calculations for this zone - strictly counting VALIDATED payments
    total_invoiced = sum(inv.amount for inv in invoices)

    # Calculate real income from Payment records associated with invoices in this zone
    zone_payments = Payment.objects.filter(invoice__client__zone=zone)
    total_paid = sum(p.amount for p in zone_payments)

    total_pending = total_invoiced - total_paid

    # Global earnings - strictly counting VALIDATED payments across all zones
    all_invoices = Invoice.objects.all()
    global_total_invoiced = sum(inv.amount for inv in all_invoices)

    all_payments = Payment.objects.all()
    global_total_paid = sum(p.amount for p in all_payments)

    # Process clients for display with subscription status and colors
    today = timezone.now().date()
    client_data = []
    for client in clients:
        sub = client.subscriptions.filter(is_active=True).first()
        status_color = "slate"  # default
        days_left = None

        if sub:
            days_left = (sub.end_date - today).days
            if days_left <= 3:
                status_color = "red"
            elif days_left <= 7:
                status_color = "orange"
            else:
                status_color = "emerald"

        # Check for pending invoices
        pending_invoices = client.invoices.filter(
            status=Invoice.Status.PENDING
        ).order_by("due_date")
        has_pending = pending_invoices.exists()
        next_invoice_due = pending_invoices.first().due_date if has_pending else None
        next_due_date = (
            calculate_next_available_invoice_due_date(client, sub, today)
            if sub
            else None
        )

        client_data.append(
            {
                "client": client,
                "subscription": sub,
                "days_left": days_left,
                "status_color": status_color,
                "has_pending": has_pending,
                "pending_amount": sum(inv.amount for inv in pending_invoices),
                "next_invoice_due": next_invoice_due,
                "next_due_date": next_due_date,
            }
        )

    context = {
        "zone": zone,
        "clients": clients,
        "client_data": client_data,
        "invoices": invoices,
        "subscriptions": subscriptions,
        "subscription_plans": subscription_plans,
        "invoice_types": Invoice.InvoiceType.choices,
        "total_invoiced": total_invoiced,
        "total_paid": total_paid,
        "total_pending": total_pending,
        "global_total_invoiced": global_total_invoiced,
        "global_total_paid": global_total_paid,
    }
    return render(request, "business/zone_finance_manage.html", context)


from django.http import JsonResponse


@login_required
def export_zone_finance_excel(request, pk):
    if not check_zone_access(request.user, pk):
        messages.error(request, "Accès refusé.")
        return redirect("dashboard")
    zone = get_object_or_404(Zone, pk=pk)

    clients = User.objects.filter(zone=zone, role=User.Role.CLIENT).prefetch_related(
        "subscriptions", "invoices"
    )
    invoices = Invoice.objects.filter(client__zone=zone).select_related("client")
    payments = Payment.objects.filter(invoice__client__zone=zone).select_related(
        "invoice__client"
    )

    total_invoiced = sum(inv.amount for inv in invoices)
    total_paid = sum(p.amount for p in payments)
    today = timezone.now().date()

    report = ExcelReportBuilder(
        f"Finance - {zone.name}",
        request=request,
        filename_prefix=f"finance_{zone.name}",
    )
    ws = report.active_sheet("Récapitulatif")
    report.add_title(ws, 12)
    start = report.add_kpis(
        ws,
        [
            ("Clients", clients.count()),
            ("Facturé", total_invoiced),
            ("Payé", total_paid),
            ("Solde", total_invoiced - total_paid),
        ],
        start_row=4,
    )

    rows = []
    for client in clients:
        sub = client.subscriptions.filter(is_active=True).first()
        client_invoices = invoices.filter(client=client)
        client_payments = payments.filter(invoice__client=client)
        client_invoiced = sum(inv.amount for inv in client_invoices)
        client_paid = sum(p.amount for p in client_payments)
        balance = client_invoiced - client_paid
        status = (
            "Actif" if sub and sub.end_date >= today else "Expiré" if sub else "Inactif"
        )
        if balance > 0:
            status += " (Dette)"
        rows.append(
            [
                client.registration_number or "-",
                client.get_full_name() or client.username,
                client.email or "-",
                client.phone_number or "-",
                client.phone_number_2 or "-",
                client.address or "-",
                sub.plan.name if sub else "Aucun",
                sub.plan.price if sub else 0,
                client_invoiced,
                client_paid,
                balance,
                status,
            ]
        )
    report.add_table(
        ws,
        [
            "Matricule",
            "Nom",
            "Email",
            "Téléphone",
            "Téléphone 2",
            "Adresse",
            "Abonnement",
            "Prix",
            "Total facturé",
            "Total payé",
            "Solde",
            "Statut",
        ],
        rows,
        start_row=start,
        table_name="FinanceClients",
    )

    ws_pay = report.create_sheet("Paiements")
    report.add_title(ws_pay, 9, "Détail des paiements de la zone")
    pay_rows = [
        [
            p.paid_at.strftime("%d/%m/%Y"),
            p.paid_at.strftime("%H:%M"),
            p.invoice.client.get_full_name() or p.invoice.client.username,
            p.invoice.client.phone_number or "-",
            p.invoice.client.registration_number or "-",
            p.invoice.get_invoice_type_display(),
            p.amount,
            p.payment_method,
            p.transaction_id,
        ]
        for p in payments.order_by("-paid_at")
    ]
    end = report.add_table(
        ws_pay,
        [
            "Date",
            "Heure",
            "Client",
            "Téléphone",
            "Matricule",
            "Type",
            "Montant",
            "Méthode",
            "Transaction",
        ],
        pay_rows,
        start_row=4,
        table_name="PaiementsZone",
    )
    report.add_total_row(ws_pay, end, 6, "TOTAL PAYÉ", {7: total_paid})
    return report.response()


from django.http import JsonResponse


@login_required
def zone_invoice_pay_all(request, pk):
    """Mark all pending invoices of a client as paid."""
    read_only_response = block_read_only_role(request)
    if read_only_response:
        return read_only_response

    if request.user.role == User.Role.ZONE_MANAGER:
        messages.error(
            request,
            "Accès refusé. Seul le comptable ou l'administrateur peut valider les paiements.",
        )
        return redirect("dashboard")

    if not check_zone_access(request.user, pk):
        messages.error(request, "Accès refusé.")
        return redirect("dashboard")
    if request.method == "POST":
        client_id = request.POST.get("client_id")
        amount_received = float(request.POST.get("amount", 0))
        payment_method = request.POST.get("payment_method", "Espèces")

        client = get_object_or_404(User, id=client_id, zone_id=pk)
        pending_invoices = client.invoices.filter(
            status=Invoice.Status.PENDING
        ).order_by("created_at")

        import uuid

        from finance.models import Payment

        remaining = amount_received
        for invoice in pending_invoices:
            if remaining <= 0:
                break

            pay_amount = min(float(invoice.amount), remaining)

            # Simple logic: if fully paid, mark as PAID. If partial, we still mark as PAID for now
            # as requested by the simplified system, but in a real system we'd track partials.
            # Here we just mark the invoice as PAID if we apply any amount to it for simplicity.
            invoice.status = Invoice.Status.PAID
            invoice.synced = False
            invoice.save()

            Payment.objects.create(
                invoice=invoice,
                amount=pay_amount,
                payment_method=payment_method,
                transaction_id=str(uuid.uuid4())[:18].upper(),
                synced=False,
            )

            if invoice.subscription:
                sub = invoice.subscription
                sub.is_active = True
                sub.synced = False
                sub.save()

            remaining -= pay_amount

        messages.success(
            request,
            f"Paiement de {amount_received} FCFA enregistré pour {client.username}.",
        )
    return redirect("zone_finance_manage", pk=pk)


from django.http import JsonResponse


@login_required
def quartier_create(request, pk):
    read_only_response = block_read_only_role(request)
    if read_only_response:
        return read_only_response

    if not check_zone_access(request.user, pk):
        messages.error(request, "Accès refusé.")
        return redirect("dashboard")
    zone = get_object_or_404(Zone, pk=pk)
    if request.method == "POST":
        name = request.POST.get("name")
        if name:
            from .models import Quartier

            Quartier.objects.create(name=name, zone=zone)
            messages.success(request, f"Quartier '{name}' ajouté.")
        else:
            messages.error(request, "Le nom du quartier est requis.")
    return redirect(request.META.get("HTTP_REFERER", "zone_clients_manage"))


from django.http import JsonResponse


@login_required
def zone_client_create(request, pk):
    read_only_response = block_read_only_role(request)
    if read_only_response:
        return read_only_response

    if request.user.role not in [
        User.Role.SUPER_ADMIN,
        User.Role.ZONE_MANAGER,
    ]:
        messages.error(
            request,
            "Accès refusé. Seul l'administrateur ou le chef de zone peut créer des clients.",
        )
        return redirect("zone_clients_manage", pk=pk)

    if not check_zone_access(request.user, pk):
        messages.error(request, "Accès refusé.")
        return redirect("dashboard")
    zone = get_object_or_404(Zone, pk=pk)
    if request.method == "POST":
        # Check for expired subscriptions first
        expired_count = check_and_deactivate_expired_subscriptions()
        expired_momo_count = check_and_deactivate_expired_momo_subscriptions()

        total_expired = expired_count + expired_momo_count
        if total_expired > 0:
            messages.info(
                request,
                f"{total_expired} abonnement(s) expiré(s) ont été désactivés automatiquement.",
            )

        username = request.POST.get("username")
        email = request.POST.get("email")
        phone = request.POST.get("phone")
        phone_2 = request.POST.get("phone_2")
        quartier_id = request.POST.get("quartier")
        address = request.POST.get("address")
        plan_id = request.POST.get("plan")

        if username:
            import random

            # Generate a unique matricule before creation to use as username
            while True:
                matricule = f"{random.randint(100000, 999999)}"
                if (
                    not User.objects.filter(username=matricule).exists()
                    and not User.objects.filter(registration_number=matricule).exists()
                ):
                    break

            from .models import Quartier

            quartier = None
            if quartier_id:
                quartier = get_object_or_404(Quartier, id=quartier_id, zone=zone)

            due_date_str = request.POST.get("due_date")

            client = User.objects.create(
                username=matricule,  # Use matricule as unique username
                first_name=username,  # Store full name in first_name
                registration_number=matricule,
                email=email,
                phone_number=phone,
                phone_number_2=phone_2,
                quartier=quartier,
                address=address,
                zone=zone,
                role=User.Role.CLIENT,
                is_active=True,  # Active account
            )

            if due_date_str:
                client.fixed_due_date = timezone.datetime.strptime(
                    due_date_str, "%Y-%m-%d"
                ).date()
            client.save()

            # Add subscription if a plan was selected
            if plan_id:
                plan = get_object_or_404(SubscriptionPlan, id=plan_id)

                if client.fixed_due_date:
                    end_date = client.fixed_due_date
                else:
                    temp_subscription = Subscription(client=client, plan=plan)
                    end_date = next_period_end_date(
                        client,
                        temp_subscription,
                        timezone.now().date(),
                    )
                    client.fixed_due_date = end_date
                    client.save(
                        update_fields=["fixed_due_date", "synced", "updated_at"]
                    )

                subscription = Subscription.objects.create(
                    client=client,
                    plan=plan,
                    end_date=end_date,
                    is_active=False,  # Explicitly inactive until payment
                )

                # Auto-generate a pending invoice for the initial payment
                from finance.models import Invoice

                Invoice.objects.create(
                    client=client,
                    subscription=subscription,
                    amount=plan.price,
                    invoice_type=Invoice.InvoiceType.PAIEMENT,
                    due_date=subscription.end_date,  # Use the subscription's fixed due date
                    synced=False,
                )

                messages.success(
                    request,
                    f"Client '{username}' créé. L'abonnement {plan.name} sera activé après le paiement de la facture générée.",
                )
            else:
                messages.success(request, f"Client '{username}' ajouté.")
        else:
            messages.error(request, "Le nom d'utilisateur est requis.")
    return redirect("zone_clients_manage", pk=pk)


from django.http import JsonResponse


@login_required
def zone_client_edit(request, zone_pk, client_pk):
    read_only_response = block_read_only_role(request)
    if read_only_response:
        return read_only_response

    if not check_zone_access(request.user, zone_pk):
        messages.error(request, "Accès refusé.")
        return redirect("dashboard")
    zone = get_object_or_404(Zone, pk=zone_pk)
    client = get_object_or_404(User, pk=client_pk, zone=zone)
    if request.method == "POST":
        new_name = request.POST.get(
            "username"
        )  # The form field is still named 'username' but it represents the full name
        if new_name:
            client.first_name = new_name

        client.email = request.POST.get("email")
        client.phone_number = request.POST.get("phone")
        client.phone_number_2 = request.POST.get("phone_2")

        due_date_str = request.POST.get("due_date")

        if due_date_str:
            client.fixed_due_date = timezone.datetime.strptime(
                due_date_str, "%Y-%m-%d"
            ).date()

        quartier_id = request.POST.get("quartier")
        if quartier_id:
            from .models import Quartier

            client.quartier = get_object_or_404(Quartier, id=quartier_id, zone=zone)
        else:
            client.quartier = None

        client.address = request.POST.get("address")
        client.save()

        # Subscription plan modification
        plan_id = request.POST.get("plan")
        if plan_id:
            plan = get_object_or_404(SubscriptionPlan, id=plan_id)
            subscription = client.subscriptions.first()

            if subscription:
                subscription.plan = plan
                subscription.end_date = client.fixed_due_date or next_period_end_date(
                    client,
                    subscription,
                    timezone.now().date(),
                )
                subscription.synced = False
                subscription.save()
            else:
                temp_subscription = Subscription(client=client, plan=plan)
                final_end_date = client.fixed_due_date or next_period_end_date(
                    client,
                    temp_subscription,
                    timezone.now().date(),
                )
                if not client.fixed_due_date:
                    client.fixed_due_date = final_end_date
                    client.save(
                        update_fields=["fixed_due_date", "synced", "updated_at"]
                    )
                Subscription.objects.create(
                    client=client, plan=plan, end_date=final_end_date
                )
            messages.success(
                request, f"Client '{client.username}' et abonnement mis à jour."
            )
        else:
            messages.success(request, f"Client '{client.username}' mis à jour.")

        return redirect("zone_clients_manage", pk=zone_pk)

    subscription_plans = SubscriptionPlan.objects.all()
    current_subscription = client.subscriptions.first()
    return render(
        request,
        "business/zone_client_form.html",
        {
            "zone": zone,
            "client": client,
            "is_edit": True,
            "subscription_plans": subscription_plans,
            "current_subscription": current_subscription,
        },
    )


from django.http import JsonResponse


@login_required
def zone_client_delete(request, zone_pk, client_pk):
    read_only_response = block_read_only_role(request)
    if read_only_response:
        return read_only_response

    if not check_zone_access(request.user, zone_pk):
        messages.error(request, "Accès refusé.")
        return redirect("dashboard")
    zone = get_object_or_404(Zone, pk=zone_pk)
    client = get_object_or_404(User, pk=client_pk, zone=zone)
    if request.method == "POST":
        client.delete()
        messages.success(request, "Client supprimé.")
        return redirect("zone_clients_manage", pk=zone_pk)
    return render(
        request,
        "business/zone_client_confirm_delete.html",
        {"zone": zone, "client": client},
    )


from django.http import JsonResponse


@login_required
def zone_employee_create(request, pk):
    if not check_zone_access(request.user, pk):
        messages.error(request, "Accès refusé.")
        return redirect("dashboard")
    zone = get_object_or_404(Zone, pk=pk)
    if request.method == "POST":
        first_name = request.POST.get("first_name")
        last_name = request.POST.get("last_name")
        phone = request.POST.get("phone")
        phone_2 = request.POST.get("phone_2")
        address = request.POST.get("address")
        role = request.POST.get("role")  # Changed from position to role

        # Restriction: only SUPER_ADMIN and ACCOUNTANT can set salary
        if request.user.role in [User.Role.SUPER_ADMIN, User.Role.ACCOUNTANT]:
            salary = request.POST.get("salary", 0)
        else:
            salary = 0

        hired_at = request.POST.get("hired_at")

        # Create or get position based on role
        position, created = Position.objects.get_or_create(title=role.upper())

        Employee.objects.create(
            first_name=first_name,
            last_name=last_name,
            phone_number=phone,
            phone_number_2=phone_2,
            address=address,
            zone=zone,
            salary=salary,
            position=position,
            hired_at=hired_at,
            synced=False,
        )
        messages.success(
            request, f"Employé '{first_name}' ajouté à la zone {zone.name}."
        )
    return redirect("zone_employees_manage", pk=pk)


from django.http import JsonResponse


@login_required
def zone_subscription_create(request, pk):
    if not check_zone_access(request.user, pk):
        messages.error(request, "Accès refusé.")
        return redirect("dashboard")
    zone = get_object_or_404(Zone, pk=pk)
    if request.method == "POST":
        client_id = request.POST.get("client")
        plan_id = request.POST.get("plan")

        client = get_object_or_404(User, id=client_id, zone=zone)
        plan = get_object_or_404(SubscriptionPlan, id=plan_id)

        temp_subscription = Subscription(client=client, plan=plan)
        end_date = client.fixed_due_date or next_period_end_date(
            client,
            temp_subscription,
            timezone.now().date(),
        )
        if not client.fixed_due_date:
            client.fixed_due_date = end_date
            client.save(update_fields=["fixed_due_date", "synced", "updated_at"])

        subscription = Subscription.objects.create(
            client=client, plan=plan, end_date=end_date, is_active=False
        )

        # Auto-generate invoice
        from finance.models import Invoice

        Invoice.objects.create(
            client=client,
            subscription=subscription,
            amount=plan.price,
            invoice_type=Invoice.InvoiceType.PAIEMENT,
            due_date=subscription.end_date,  # Use the subscription's fixed due date
            synced=False,
        )

        messages.success(request, "Abonnement créé en attente de paiement.")
    return redirect("zone_detail", pk=pk)


from django.http import JsonResponse


@login_required
def subscription_edit(request, pk):
    subscription = get_object_or_404(Subscription, pk=pk)
    if subscription.client.zone and not check_zone_access(
        request.user, subscription.client.zone.id
    ):
        messages.error(request, "Accès refusé.")
        return redirect("dashboard")
    zone_pk = subscription.client.zone.pk
    if request.method == "POST":
        plan_id = request.POST.get("plan")
        is_active = request.POST.get("is_active") == "on"

        plan = get_object_or_404(SubscriptionPlan, id=plan_id)
        subscription.plan = plan
        subscription.is_active = is_active
        subscription.save()
        messages.success(request, "Abonnement mis à jour.")
    return redirect("zone_detail", pk=zone_pk)


from django.http import JsonResponse


@login_required
def subscription_delete(request, pk):
    read_only_response = block_read_only_role(request)
    if read_only_response:
        return read_only_response

    subscription = get_object_or_404(Subscription, pk=pk)
    if subscription.client.zone and not check_zone_access(
        request.user, subscription.client.zone.id
    ):
        messages.error(request, "Accès refusé.")
        return redirect("dashboard")
    zone_pk = subscription.client.zone.pk
    if request.method == "POST":
        subscription.delete()
        messages.success(request, "Abonnement supprimé.")
    else:
        messages.error(request, "Action non autorisée sans confirmation.")
    return redirect("zone_detail", pk=zone_pk)


from django.http import JsonResponse


@login_required
def zone_invoice_pay(request, pk):
    """Mark an invoice as paid and record the payment."""
    read_only_response = block_read_only_role(request)
    if read_only_response:
        return read_only_response

    if request.user.role == User.Role.ZONE_MANAGER:
        messages.error(
            request,
            "Accès refusé. Seul le comptable ou l'administrateur peut valider les paiements.",
        )
        return redirect("dashboard")

    invoice = get_object_or_404(Invoice, pk=pk)
    if invoice.client.zone and not check_zone_access(
        request.user, invoice.client.zone.id
    ):
        messages.error(request, "Accès refusé.")
        return redirect("dashboard")
    zone_pk = invoice.client.zone.pk

    if request.method == "POST":
        import uuid

        from finance.models import Payment

        # Update invoice status
        invoice.status = Invoice.Status.PAID
        invoice.synced = False
        invoice.save()

        # Create Payment record
        Payment.objects.create(
            invoice=invoice,
            amount=invoice.amount,
            payment_method=request.POST.get("payment_method", "Espèces"),
            transaction_id=str(uuid.uuid4())[:18].upper(),
            synced=False,
        )

        # If this invoice was for a subscription, ensure the subscription is active
        if invoice.subscription:
            sub = invoice.subscription

            sub.end_date = payment_subscription_end_date(invoice)
            sub.is_active = True
            sub.synced = False
            sub.save()

        messages.success(
            request, f"Paiement validé pour la facture de {invoice.client.username}."
        )

    return redirect("zone_finance_manage", pk=zone_pk)


from django.http import JsonResponse


@login_required
def zone_invoice_create(request, pk):
    read_only_response = block_read_only_role(request)
    if read_only_response:
        return read_only_response

    if not check_zone_access(request.user, pk):
        messages.error(request, "Accès refusé.")
        return redirect("dashboard")
    zone = get_object_or_404(Zone, pk=pk)
    if request.method == "POST":
        client_id = request.POST.get("client")
        subscription_id = request.POST.get("subscription")
        amount = Decimal(request.POST.get("amount", "0"))
        invoice_type = request.POST.get("invoice_type", Invoice.InvoiceType.PAIEMENT)

        client = get_object_or_404(User, id=client_id, zone=zone)
        subscription = None

        if subscription_id:
            subscription = get_object_or_404(Subscription, id=subscription_id)
        else:
            subscription = client.subscriptions.filter(is_active=True).first()

        if not subscription:
            messages.error(request, "Ce client n'a pas d'abonnement actif.")
            return redirect("zone_finance_manage", pk=pk)

        plan_price = subscription.plan.price
        amount = plan_price

        calculated_due_date = calculate_next_available_invoice_due_date(
            client,
            subscription,
            timezone.now().date(),
        )

        Invoice.objects.create(
            client=client,
            subscription=subscription,
            amount=amount,
            invoice_type=invoice_type,
            due_date=calculated_due_date,
            synced=False,
        )
        messages.success(
            request,
            f"Facture de {amount} FCFA générée pour {client.first_name or client.username}.",
        )
    return redirect("zone_finance_manage", pk=pk)


from django.http import JsonResponse


@login_required
def client_card(request, pk):
    client = get_object_or_404(User, pk=pk, role=User.Role.CLIENT)
    if client.zone and not check_zone_access(request.user, client.zone.id):
        messages.error(request, "Accès refusé.")
        return redirect("dashboard")
    subscription = client.subscriptions.first()
    context = {
        "client": client,
        "subscription": subscription,
        "now": timezone.now(),
    }
    return render(request, "business/client_card.html", context)


from django.http import JsonResponse


@login_required
def invoice_print(request, pk):
    invoice = get_object_or_404(Invoice, pk=pk)

    # Permission check: Owner can view, or staff with zone access
    can_view = False
    if request.user == invoice.client:
        can_view = True
    elif request.user.role in [
        User.Role.SUPER_ADMIN,
        User.Role.ACCOUNTANT,
        User.Role.SHAREHOLDER,
    ]:
        can_view = True
    elif request.user.role == User.Role.ZONE_MANAGER:
        if invoice.client.zone and check_zone_access(
            request.user, invoice.client.zone.id
        ):
            can_view = True

    if not can_view:
        messages.error(request, "Accès refusé.")
        return redirect("dashboard")

    total_price = invoice.amount
    balance = 0
    paid_already = invoice.amount

    # Determine where to go back
    back_url = "invoice_list"
    if request.user.role == User.Role.CLIENT:
        back_url = "accounts:client_dashboard"

    return render(
        request,
        "finance/invoice_a4.html",
        {
            "invoice": invoice,
            "total_price": total_price,
            "balance": balance,
            "paid_already": paid_already,
            "now": timezone.now(),
            "back_url_name": back_url,
        },
    )


from django.http import JsonResponse


@login_required
def client_payment_history(request, pk):
    client = get_object_or_404(User, pk=pk, role=User.Role.CLIENT)
    if client.zone and not check_zone_access(request.user, client.zone.id):
        return JsonResponse({"error": "Accès refusé"}, status=403)

    invoices = Invoice.objects.filter(client=client).order_by("-due_date")

    # Building a "Calendar" view for the current year
    now = timezone.now()
    year = now.year
    months_data = []
    month_names = [
        "Jan",
        "Fév",
        "Mar",
        "Avr",
        "Mai",
        "Juin",
        "Juil",
        "Août",
        "Sep",
        "Oct",
        "Nov",
        "Déc",
    ]

    for i in range(1, 13):
        # Check if an invoice exists for this month and if it's paid
        month_invoice = invoices.filter(due_date__month=i, due_date__year=year).first()
        status = "none"  # No invoice
        if month_invoice:
            status = (
                "paid" if month_invoice.status == Invoice.Status.PAID else "pending"
            )

        months_data.append(
            {
                "name": month_names[i - 1],
                "status": status,
                "invoice_id": month_invoice.id if month_invoice else None,
            }
        )

    payment_list = []
    for inv in invoices:
        payment_list.append(
            {
                "date": inv.due_date.strftime("%d/%m/%Y"),
                "amount": f"{inv.amount} FCFA",
                "type": inv.get_invoice_type_display(),
                "status": inv.get_status_display(),
                "status_code": inv.status,
            }
        )

    subscription = client.subscriptions.first()

    return JsonResponse(
        {
            "client_name": client.get_full_name() or client.username,
            "registration_number": client.registration_number,
            "months": months_data,
            "payments": payment_list,
            "subscription": {
                "id": subscription.id if subscription else None,
                "is_active": subscription.is_active if subscription else False,
                "plan": subscription.plan.name if subscription else "Aucun",
            },
        }
    )


from django.http import JsonResponse


@login_required
def subscription_toggle_status(request, pk):
    read_only_response = block_read_only_role(request)
    if read_only_response:
        return read_only_response

    subscription = get_object_or_404(Subscription, pk=pk)
    if subscription.client.zone and not check_zone_access(
        request.user, subscription.client.zone.id
    ):
        messages.error(request, "Accès refusé.")
        return redirect("dashboard")

    subscription.is_active = not subscription.is_active
    subscription.synced = False
    subscription.save()

    status_msg = "activé" if subscription.is_active else "désactivé"
    messages.success(
        request, f"L'abonnement de {subscription.client.username} a été {status_msg}."
    )
    return redirect(request.META.get("HTTP_REFERER", "dashboard"))
